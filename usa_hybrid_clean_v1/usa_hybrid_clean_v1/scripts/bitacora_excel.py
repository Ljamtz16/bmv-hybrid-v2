"""
Bitácora Excel de predicciones H3 con monitoreo de precios en tiempo real.
Registra todas las predicciones y actualiza su progreso automáticamente.
"""
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Ruta de la bitácora - configurada para Google Drive
# Si la ruta de Drive no existe, usa la local como fallback
DRIVE_PATH = r"G:\Mi unidad\Trading proyecto\H3_BITACORA_PREDICCIONES.xlsx"
LOCAL_PATH = "reports/H3_BITACORA_PREDICCIONES.xlsx"

if os.path.exists(os.path.dirname(DRIVE_PATH)):
    BITACORA_PATH = DRIVE_PATH
    print(f"[DRIVE] Usando Google Drive: {BITACORA_PATH}")
else:
    BITACORA_PATH = LOCAL_PATH
    print(f"[LOCAL] Usando ruta local (Drive no disponible): {BITACORA_PATH}")

def init_bitacora():
    """Crear bitácora si no existe con formato profesional."""
    if os.path.exists(BITACORA_PATH):
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Predicciones"
    
    # Headers
    headers = [
        "ID", "Fecha Predicción", "Ticker", "Side", "Entry Price", 
        "TP Price", "SL Price", "TP %", "SL %",
        "Prob Win", "Y_hat", "Horizon (días)", "ETTH (días)", "P(TP≺SL)",
        "Score TTH", "Sector",
        "Status", "Fecha Cierre", "Exit Price", "PnL USD", "PnL %", "Días Transcurridos",
        "Precio Actual", "Última Actualización", "Progreso a TP %", "Notas"
    ]
    
    ws.append(headers)
    
    # Estilos para headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajustar anchos
    widths = [35, 18, 10, 8, 12, 12, 12, 8, 8, 10, 10, 12, 12, 12, 10, 12, 12, 18, 12, 12, 10, 15, 12, 20, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    wb.save(BITACORA_PATH)
    print(f"[OK] Bitacora creada: {BITACORA_PATH}")

def add_prediction(trade_row, forecast_date=None):
    """
    Añadir nueva predicción a la bitácora.
    
    Args:
        trade_row: dict o pandas.Series con los datos del trade
        forecast_date: fecha del forecast (default: hoy)
    """
    init_bitacora()
    
    if forecast_date is None:
        forecast_date = datetime.now()
    
    # Generar ID único
    prediction_id = f"{trade_row['ticker']}_{forecast_date.strftime('%Y%m%d_%H%M%S')}"
    
    # Calcular precios TP y SL según side
    entry = float(trade_row.get('entry_price', 0))
    tp_pct = float(trade_row.get('tp_pct', 0.04))
    sl_pct = float(trade_row.get('sl_pct', 0.02))
    side = trade_row.get('side', 'BUY')
    
    if side == 'BUY':
        tp_price = entry * (1 + tp_pct)
        sl_price = entry * (1 - sl_pct)
    else:  # SHORT
        tp_price = entry * (1 - tp_pct)  # TP abajo
        sl_price = entry * (1 + sl_pct)  # SL arriba
    
    # Preparar datos
    new_row = {
        "ID": prediction_id,
        "Fecha Predicción": forecast_date.strftime("%Y-%m-%d %H:%M"),
        "Ticker": trade_row.get('ticker', ''),
        "Side": trade_row.get('side', 'BUY'),
        "Entry Price": entry,
        "TP Price": tp_price,
        "SL Price": sl_price,
        "TP %": tp_pct * 100,
        "SL %": sl_pct * 100,
        "Prob Win": float(trade_row.get('prob_win', 0)),
        "Y_hat": float(trade_row.get('y_hat', 0)),
        "Horizon (días)": int(trade_row.get('horizon_days', 3)),
        "ETTH (días)": float(trade_row.get('etth_first_event', 0)),
        "P(TP≺SL)": float(trade_row.get('p_tp_before_sl', 0)),
        "Score TTH": float(trade_row.get('tth_score', 0)),
        "Sector": trade_row.get('sector', 'Unknown'),
        "Status": "ACTIVO",
        "Fecha Cierre": "",
        "Exit Price": "",
        "PnL USD": "",
        "PnL %": "",
        "Días Transcurridos": 0,
        "Precio Actual": entry,
        "Última Actualización": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Progreso a TP %": 0.0,
        "Notas": trade_row.get('notes', '')
    }
    
    # Leer bitácora existente
    df = pd.read_excel(BITACORA_PATH, sheet_name="Predicciones")

    # Anti-duplicado: mismo Ticker + Fecha Predicción (sin minutos) + Side
    try:
        forecast_day = pd.to_datetime(new_row["Fecha Predicción"], errors="coerce").date()
        pred_dt = pd.to_datetime(df["Fecha Predicción"], errors="coerce")
        same_day = pred_dt.dt.date == forecast_day
        exists_mask = (
            (df["Ticker"] == new_row["Ticker"]) &
            (df["Side"] == new_row["Side"]) &
            same_day
        )
        if exists_mask.any():
            print(f"[SKIP] Duplicado detectado para {new_row['Ticker']} {forecast_day} {new_row['Side']} → saltando")
            return None
    except Exception:
        pass

    # Añadir nueva fila
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    
    # Guardar
    with pd.ExcelWriter(BITACORA_PATH, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name="Predicciones", index=False)
    
    print(f"[OK] Prediccion anadida: {prediction_id}")
    return prediction_id

def update_prices_from_daily(daily_csv_path="data/us/ohlcv_us_daily.csv"):
    """
    Actualizar precios actuales de todas las predicciones activas.
    
    Args:
        daily_csv_path: Ruta al CSV con precios diarios actualizados
    """
    init_bitacora()
    
    if not os.path.exists(daily_csv_path):
        print(f"[WARN] No se encuentra {daily_csv_path}")
        return
    
    # Cargar precios actuales
    prices = pd.read_csv(daily_csv_path)
    prices['date'] = pd.to_datetime(prices['date'])
    latest_prices = prices.groupby('ticker').tail(1)[['ticker', 'date', 'close']].set_index('ticker')
    
    # Mostrar fecha de precios
    latest_date = prices['date'].max().strftime("%Y-%m-%d")
    print(f"[INFO] Fecha de precios: {latest_date}")
    
    # Cargar bitácora
    df = pd.read_excel(BITACORA_PATH, sheet_name="Predicciones")
    
    # Actualizar solo predicciones activas
    activas = df['Status'] == 'ACTIVO'
    updated_count = 0
    tp_hits = 0
    sl_hits = 0
    
    for idx in df[activas].index:
        ticker = df.at[idx, 'Ticker']
        
        if ticker not in latest_prices.index:
            continue
        
        current_price = float(latest_prices.loc[ticker, 'close'])
        current_date = latest_prices.loc[ticker, 'date']
        entry_price = float(df.at[idx, 'Entry Price'])
        tp_price = float(df.at[idx, 'TP Price'])
        sl_price = float(df.at[idx, 'SL Price'])
        side = df.at[idx, 'Side']
        
        # Actualizar precio actual y fecha
        df.at[idx, 'Precio Actual'] = current_price
        df.at[idx, 'Última Actualización'] = current_date.strftime("%Y-%m-%d %H:%M")
        
        # Calcular progreso a TP según side
        if side == 'BUY':
            # LONG: TP arriba
            if tp_price > entry_price:
                progress = ((current_price - entry_price) / (tp_price - entry_price)) * 100
            else:
                progress = 0
        else:  # SHORT
            # SHORT: TP abajo (invertido)
            if tp_price < entry_price:
                progress = ((entry_price - current_price) / (entry_price - tp_price)) * 100
            else:
                progress = 0
        
        df.at[idx, 'Progreso a TP %'] = round(progress, 2)
        
        # Calcular días transcurridos
        fecha_pred = pd.to_datetime(df.at[idx, 'Fecha Predicción'])
        dias = (datetime.now() - fecha_pred).days
        df.at[idx, 'Días Transcurridos'] = dias
        
        # Verificar si alcanzó TP o SL
        horizon = int(df.at[idx, 'Horizon (días)']) if pd.notna(df.at[idx, 'Horizon (días)']) else 3
        
        if side == 'BUY':
            if current_price >= tp_price:
                df.at[idx, 'Status'] = 'TP_HIT'
                df.at[idx, 'Fecha Cierre'] = current_date.strftime("%Y-%m-%d")
                df.at[idx, 'Exit Price'] = tp_price
                pnl_usd = (tp_price - entry_price) * 100  # Asumir 100 acciones
                pnl_pct = ((tp_price - entry_price) / entry_price) * 100
                df.at[idx, 'PnL USD'] = round(pnl_usd, 2)
                df.at[idx, 'PnL %'] = round(pnl_pct, 2)
                tp_hits += 1
                print(f"  [TP] {ticker} TP HIT @ ${current_price:.2f} (+{pnl_pct:.2f}%)")
            elif current_price <= sl_price:
                df.at[idx, 'Status'] = 'SL_HIT'
                df.at[idx, 'Fecha Cierre'] = current_date.strftime("%Y-%m-%d")
                df.at[idx, 'Exit Price'] = sl_price
                pnl_usd = (sl_price - entry_price) * 100
                pnl_pct = ((sl_price - entry_price) / entry_price) * 100
                df.at[idx, 'PnL USD'] = round(pnl_usd, 2)
                df.at[idx, 'PnL %'] = round(pnl_pct, 2)
                sl_hits += 1
                print(f"  [SL] {ticker} SL HIT @ ${current_price:.2f} ({pnl_pct:.2f}%)")
        else:  # SHORT
            if current_price <= tp_price:
                df.at[idx, 'Status'] = 'TP_HIT'
                df.at[idx, 'Fecha Cierre'] = current_date.strftime("%Y-%m-%d")
                df.at[idx, 'Exit Price'] = tp_price
                pnl_usd = (entry_price - tp_price) * 100  # Invertido
                pnl_pct = ((entry_price - tp_price) / entry_price) * 100
                df.at[idx, 'PnL USD'] = round(pnl_usd, 2)
                df.at[idx, 'PnL %'] = round(pnl_pct, 2)
                tp_hits += 1
                print(f"  [TP] {ticker} SHORT TP HIT @ ${current_price:.2f} (+{pnl_pct:.2f}%)")
            elif current_price >= sl_price:
                df.at[idx, 'Status'] = 'SL_HIT'
                df.at[idx, 'Fecha Cierre'] = current_date.strftime("%Y-%m-%d")
                df.at[idx, 'Exit Price'] = sl_price
                pnl_usd = (entry_price - sl_price) * 100  # Invertido
                pnl_pct = ((entry_price - sl_price) / entry_price) * 100
                df.at[idx, 'PnL USD'] = round(pnl_usd, 2)
                df.at[idx, 'PnL %'] = round(pnl_pct, 2)
                sl_hits += 1
                print(f"  [SL] {ticker} SHORT SL HIT @ ${current_price:.2f} ({pnl_pct:.2f}%)")
        
        # Verificar expiración
        if dias >= horizon and df.at[idx, 'Status'] == 'ACTIVO':
            df.at[idx, 'Status'] = 'EXPIRED'
            df.at[idx, 'Fecha Cierre'] = current_date.strftime("%Y-%m-%d")
            df.at[idx, 'Exit Price'] = current_price
            if side == 'BUY':
                pnl_usd = (current_price - entry_price) * 100
                pnl_pct = ((current_price - entry_price) / entry_price) * 100
            else:
                pnl_usd = (entry_price - current_price) * 100
                pnl_pct = ((entry_price - current_price) / entry_price) * 100
            df.at[idx, 'PnL USD'] = round(pnl_usd, 2)
            df.at[idx, 'PnL %'] = round(pnl_pct, 2)
            print(f"  [EXP] {ticker} EXPIRED despues de {dias} dias ({pnl_pct:+.2f}%)")
        
        updated_count += 1
    
    # Guardar con formato
    with pd.ExcelWriter(BITACORA_PATH, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name="Predicciones", index=False)
    
    # Aplicar formato condicional
    _apply_conditional_formatting(BITACORA_PATH)
    
    # Mostrar resumen
    print(f"[OK] Actualizadas {updated_count} predicciones")
    if tp_hits > 0:
        print(f"   [TP] {tp_hits} alcanzaron TP")
    if sl_hits > 0:
        print(f"   [SL] {sl_hits} alcanzaron SL")
    
    return updated_count

def _apply_conditional_formatting(path):
    """Aplicar formato condicional a la bitácora."""
    wb = load_workbook(path)
    ws = wb["Predicciones"]
    
    # Definir estilos
    tp_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    sl_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    active_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Encontrar columna Status
    status_col = None
    for cell in ws[1]:
        if cell.value == "Status":
            status_col = cell.column
            break
    
    if status_col:
        for row in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row, column=status_col)
            status = status_cell.value
            
            # Colorear toda la fila según status
            if status == "TP_HIT":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = tp_fill
            elif status == "SL_HIT":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = sl_fill
            elif status == "ACTIVO":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = active_fill
    
    wb.save(path)

def generate_summary_sheet():
    """Generar hoja de resumen con estadísticas."""
    init_bitacora()
    
    df = pd.read_excel(BITACORA_PATH, sheet_name="Predicciones")
    
    # Calcular estadísticas
    total = len(df)
    activos = len(df[df['Status'] == 'ACTIVO'])
    tp_hit = len(df[df['Status'] == 'TP_HIT'])
    sl_hit = len(df[df['Status'] == 'SL_HIT'])
    
    cerrados = tp_hit + sl_hit
    win_rate = (tp_hit / cerrados * 100) if cerrados > 0 else 0
    
    # PnL total
    df_closed = df[df['Status'].isin(['TP_HIT', 'SL_HIT'])]
    total_pnl = df_closed['PnL USD'].sum() if 'PnL USD' in df_closed.columns and len(df_closed) > 0 else 0
    
    summary = {
        "Métrica": [
            "Total Predicciones",
            "Activas",
            "TP Hit",
            "SL Hit",
            "Win Rate %",
            "PnL Total USD",
            "Última Actualización"
        ],
        "Valor": [
            total,
            activos,
            tp_hit,
            sl_hit,
            round(win_rate, 2),
            round(total_pnl, 2),
            datetime.now().strftime("%Y-%m-%d %H:%M")
        ]
    }
    
    summary_df = pd.DataFrame(summary)
    
    # Guardar en nueva hoja
    with pd.ExcelWriter(BITACORA_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        summary_df.to_excel(writer, sheet_name="Resumen", index=False)
    
    print(f"✅ Hoja de resumen actualizada")
    print(f"   Total: {total} | Activas: {activos} | Win Rate: {win_rate:.1f}%")

def add_trades_from_plan(plan_csv_path, forecast_date=None):
    """
    Añadir todas las predicciones de un trade plan a la bitácora.
    
    Args:
        plan_csv_path: Ruta al CSV del trade plan (ej: trade_plan_tth.csv)
    """
    if not os.path.exists(plan_csv_path):
        print(f"⚠️  No se encuentra {plan_csv_path}")
        return
    
    plan = pd.read_csv(plan_csv_path)
    
    added = 0
    for _, row in plan.iterrows():
        try:
            add_prediction(row, forecast_date=forecast_date)
            added += 1
        except Exception as e:
            print(f"[WARN] Error anadiendo {row.get('ticker', 'unknown')}: {e}")
    
    print(f"[OK] {added} predicciones anadidas desde {plan_csv_path}")
    generate_summary_sheet()

if __name__ == "__main__":
    import argparse
    
    ap = argparse.ArgumentParser()
    ap.add_argument("--init", action="store_true", help="Crear bitácora vacía")
    ap.add_argument("--add-plan", default="", help="Añadir predicciones desde trade_plan_tth.csv")
    ap.add_argument("--forecast-date", default="", help="Fecha de la predicción (YYYY-MM-DD o YYYY-MM-DD HH:MM)")
    ap.add_argument("--update", action="store_true", help="Actualizar precios de predicciones activas")
    ap.add_argument("--summary", action="store_true", help="Generar hoja de resumen")
    args = ap.parse_args()
    
    if args.init:
        init_bitacora()
    elif args.add_plan:
        # Parse optional forecast date
        fdate = None
        if args.forecast_date:
            try:
                # Try full datetime first
                fdate = datetime.strptime(args.forecast_date, "%Y-%m-%d %H:%M")
            except ValueError:
                try:
                    fdate = datetime.strptime(args.forecast_date, "%Y-%m-%d")
                except ValueError:
                    print(f"[WARN] Formato de fecha no reconocido: {args.forecast_date}. Usando ahora().")
                    fdate = None
        add_trades_from_plan(args.add_plan, forecast_date=fdate)
    elif args.update:
        update_prices_from_daily()
        generate_summary_sheet()
    elif args.summary:
        generate_summary_sheet()
    else:
        print("Uso:")
        print("  python scripts/bitacora_excel.py --init")
        print("  python scripts/bitacora_excel.py --add-plan reports/forecast/2025-11/trade_plan_tth.csv")
        print("  python scripts/bitacora_excel.py --update")
        print("  python scripts/bitacora_excel.py --summary")
