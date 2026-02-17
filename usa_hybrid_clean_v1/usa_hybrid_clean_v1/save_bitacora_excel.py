#!/usr/bin/env python3
"""
Script para guardar todos los planes y trades en el archivo Excel de bitácora.
Ubicación: G:\Mi unidad\Trading proyecto\H3_BITACORA_PREDICCIONES.xlsx
"""
import os
from datetime import datetime
from pathlib import Path
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración
PLANS_DIR = Path("evidence/weekly_plans")
WEEK_START = "2026-01-26"
TRADE_PLAN_PATH = Path("val/trade_plan_EXECUTE.csv")
EXCEL_PATH = r"G:\Mi unidad\Trading proyecto\H3_BITACORA_PREDICCIONES.xlsx"

# Crear directorio si no existe
os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)

print("="*80)
print("GUARDANDO REGISTRO DE PLANES Y TRADES EN EXCEL")
print("="*80)

# Cargar planes
print("\n[1] Cargando planes...")
std_file = PLANS_DIR / f"plan_standard_{WEEK_START}.csv"
pw_file = PLANS_DIR / f"plan_probwin55_{WEEK_START}.csv"

df_standard = pd.read_csv(std_file) if std_file.exists() else pd.DataFrame()
df_probwin = pd.read_csv(pw_file) if pw_file.exists() else pd.DataFrame()

print(f"  ✓ STANDARD: {len(df_standard)} posiciones")
print(f"  ✓ PROBWIN_55: {len(df_probwin)} posiciones")

# Cargar trades ejecutados
print("\n[2] Cargando trades ejecutados...")
df_execute = pd.read_csv(TRADE_PLAN_PATH) if TRADE_PLAN_PATH.exists() else pd.DataFrame()
print(f"  ✓ EXECUTE: {len(df_execute)} trades")

# Obtener precios actuales
print("\n[3] Obteniendo precios actuales...")
tickers = set()
for df in [df_standard, df_probwin, df_execute]:
    if not df.empty:
        tickers.update(df['ticker'].unique())

prices = {}
for ticker in tickers:
    try:
        tk = yf.Ticker(ticker)
        hist = tk.history(period="1d", interval="1m")
        if not hist.empty:
            price = float(hist['Close'].iloc[-1])
            prices[ticker] = price
            print(f"  • {ticker}: ${price:.2f}")
    except:
        prices[ticker] = None

# Función para enriquecer datos con P&L
def enriquecer_trades(df, prices, plan_type):
    """Añade precios actuales y calcula P&L"""
    df_enr = df.copy()
    df_enr['current_price'] = df_enr['ticker'].map(prices)
    df_enr['plan_type'] = plan_type
    df_enr['timestamp'] = datetime.now().isoformat()
    
    # Calcular P&L
    df_enr['pnl_unit'] = df_enr.apply(
        lambda r: r['current_price'] - r['entry'] if pd.notna(r['current_price']) else None,
        axis=1
    )
    df_enr['pnl_total'] = df_enr['pnl_unit'] * df_enr['qty']
    df_enr['pnl_pct'] = df_enr.apply(
        lambda r: (r['pnl_unit'] / r['entry'] * 100) if pd.notna(r['pnl_unit']) and r['entry'] else None,
        axis=1
    )
    
    return df_enr

# Enriquecer datos
print("\n[4] Calculando P&L...")
df_std_enr = enriquecer_trades(df_standard, prices, 'STANDARD')
df_pw_enr = enriquecer_trades(df_probwin, prices, 'PROBWIN_55')
df_exec_enr = enriquecer_trades(df_execute, prices, 'EXECUTE')

print("  ✓ P&L calculado")

# Crear o cargar workbook
print("\n[5] Escribiendo en Excel...")
if os.path.exists(EXCEL_PATH):
    wb = load_workbook(EXCEL_PATH)
    # Remover hojas antiguas si existen
    for sheet in ['STANDARD', 'PROBWIN_55', 'EXECUTE', 'RESUMEN']:
        if sheet in wb.sheetnames:
            del wb[sheet]
else:
    from openpyxl import Workbook
    wb = Workbook()
    del wb['Sheet']

# Estilos
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def estilizar_hoja(ws, df):
    """Aplica estilos a la hoja"""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if cell.row == 1:  # Header
                cell.fill = header_fill
                cell.font = header_font
            elif cell.column in [14, 15, 16]:  # P&L columns
                if cell.value and isinstance(cell.value, (int, float)):
                    if cell.value > 0:
                        cell.font = Font(color="008000", bold=True)
                    elif cell.value < 0:
                        cell.font = Font(color="FF0000", bold=True)
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

# Escribir STANDARD
ws_std = wb.create_sheet('STANDARD', 0)
for r_idx, row in enumerate(dataframe_to_rows(df_std_enr, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws_std.cell(row=r_idx, column=c_idx, value=value)
estilizar_hoja(ws_std, df_std_enr)
print("  ✓ Hoja STANDARD creada")

# Escribir PROBWIN_55
ws_pw = wb.create_sheet('PROBWIN_55', 1)
for r_idx, row in enumerate(dataframe_to_rows(df_pw_enr, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws_pw.cell(row=r_idx, column=c_idx, value=value)
estilizar_hoja(ws_pw, df_pw_enr)
print("  ✓ Hoja PROBWIN_55 creada")

# Escribir EXECUTE
ws_exec = wb.create_sheet('EXECUTE', 2)
for r_idx, row in enumerate(dataframe_to_rows(df_exec_enr, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws_exec.cell(row=r_idx, column=c_idx, value=value)
estilizar_hoja(ws_exec, df_exec_enr)
print("  ✓ Hoja EXECUTE creada")

# Crear hoja RESUMEN
print("\n[6] Creando resumen...")
ws_resumen = wb.create_sheet('RESUMEN', 3)

summary_data = [
    ['FECHA GENERACIÓN', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ['SEMANA', WEEK_START],
    ['', ''],
    ['PLAN', 'Posiciones', 'Exposición Total', 'P&L Total', 'P&L %', 'Avg Prob Win'],
]

# Calcular resumen STANDARD
std_exp = df_std_enr['exposure'].sum()
std_pnl = df_std_enr['pnl_total'].sum()
std_pnl_pct = (std_pnl / std_exp * 100) if std_exp else 0
std_prob = df_std_enr['prob_win'].mean() * 100 if len(df_std_enr) > 0 else 0
summary_data.append(['STANDARD', len(df_std_enr), f"${std_exp:.2f}", f"${std_pnl:.2f}", f"{std_pnl_pct:.2f}%", f"{std_prob:.2f}%"])

# Calcular resumen PROBWIN_55
pw_exp = df_pw_enr['exposure'].sum()
pw_pnl = df_pw_enr['pnl_total'].sum()
pw_pnl_pct = (pw_pnl / pw_exp * 100) if pw_exp else 0
pw_prob = df_pw_enr['prob_win'].mean() * 100 if len(df_pw_enr) > 0 else 0
summary_data.append(['PROBWIN_55', len(df_pw_enr), f"${pw_exp:.2f}", f"${pw_pnl:.2f}", f"{pw_pnl_pct:.2f}%", f"{pw_prob:.2f}%"])

# Calcular resumen EXECUTE
exec_exp = df_exec_enr['exposure'].sum()
exec_pnl = df_exec_enr['pnl_total'].sum()
exec_pnl_pct = (exec_pnl / exec_exp * 100) if exec_exp else 0
exec_prob = df_exec_enr['prob_win'].mean() * 100 if len(df_exec_enr) > 0 else 0
summary_data.append(['EXECUTE', len(df_exec_enr), f"${exec_exp:.2f}", f"${exec_pnl:.2f}", f"{exec_pnl_pct:.2f}%", f"{exec_prob:.2f}%"])

# Escribir resumen
for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_resumen.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if row_idx == 4:  # Header del resumen
            cell.fill = header_fill
            cell.font = header_font
        elif row_idx > 4:  # Datos
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Ajustar ancho de columnas en RESUMEN
for col in range(1, 7):
    ws_resumen.column_dimensions[chr(64+col)].width = 20

print("  ✓ Hoja RESUMEN creada")

# Guardar
wb.save(EXCEL_PATH)
print(f"\n✅ BITÁCORA GUARDADA EN: {EXCEL_PATH}")

# Mostrar resumen final
print("\n" + "="*80)
print("RESUMEN DE REGISTROS")
print("="*80)
print(f"\nESTANDAR (Prob Win >= 0.50)")
print(f"  • Posiciones: {len(df_std_enr)}")
print(f"  • Exposición: ${std_exp:.2f}")
print(f"  • P&L Total: ${std_pnl:.2f} ({std_pnl_pct:.2f}%)")
print(f"  • Prob Win Promedio: {std_prob:.2f}%")

print(f"\nPROBWIN_55 (Prob Win >= 0.55)")
print(f"  • Posiciones: {len(df_pw_enr)}")
print(f"  • Exposición: ${pw_exp:.2f}")
print(f"  • P&L Total: ${pw_pnl:.2f} ({pw_pnl_pct:.2f}%)")
print(f"  • Prob Win Promedio: {pw_prob:.2f}%")

print(f"\nEXECUTE (Trades en seguimiento)")
print(f"  • Posiciones: {len(df_exec_enr)}")
print(f"  • Exposición: ${exec_exp:.2f}")
print(f"  • P&L Total: ${exec_pnl:.2f} ({exec_pnl_pct:.2f}%)")
print(f"  • Prob Win Promedio: {exec_prob:.2f}%")

print("\n✓ Proceso completado exitosamente")
