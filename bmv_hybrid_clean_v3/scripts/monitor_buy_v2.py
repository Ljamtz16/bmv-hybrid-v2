# monitor_bmv_buy_atr_rr_hybrid_excel_v5.py
import os, json, time, argparse
from datetime import datetime, timedelta, date
from typing import Dict, Tuple, List
import pandas as pd
import numpy as np
import yfinance as yf
import requests
from dotenv import load_dotenv, find_dotenv
from openpyxl import Workbook, load_workbook
from collections import Counter

# ========= Telegram =========
load_dotenv()
env_path = find_dotenv(usecwd=True)
loaded = load_dotenv(env_path, override=True)
print("ENV loaded:", loaded, "| path:", env_path or "<none>")

TOKEN = (os.getenv("TELEGRAM_BOT_TOKEN") or "").strip()
CHAT_ID = (os.getenv("TELEGRAM_CHAT_ID") or "").strip()

def send_tg(msg: str) -> bool:
    if not TOKEN or not CHAT_ID:
        print(f"[TG] Faltan credenciales: TOKEN={'OK' if TOKEN else 'MISSING'} CHAT_ID={CHAT_ID}")
        return False
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{TOKEN}/sendMessage",
            data={"chat_id": CHAT_ID, "text": msg, "parse_mode": "HTML"},
            timeout=15,
        )
        ok = r.status_code == 200 and r.json().get("ok", False)
        if not ok:
            print(f"[TG] Error: HTTP {r.status_code} {r.text}")
        return ok
    except Exception as e:
        print(f"[TG] Excepción: {e}")
        return False

# ========= Utilidades =========
def read_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def safe_float(x, default=None):
    try:
        if x is None: return default
        if isinstance(x, (int, float)): return float(x)
        return float(str(x).replace(",", ""))
    except:
        return default

def ensure_mx(ticker: str) -> str:
    t = ticker.strip()
    return t if t.endswith(".MX") else f"{t}.MX"

def latest_business_day_in(df_dates: pd.Series, until: date) -> pd.Timestamp | None:
    d = pd.to_datetime(df_dates).dt.date
    d = d[d <= until]
    if d.empty: return None
    return pd.Timestamp(max(d))

def pretty_cash(x: float) -> str:
    try:
        return f"${float(x):,.2f}"
    except:
        return str(x)

def prob_to_pct(prob):
    try:
        p = float(prob)
        if 0.0 <= p <= 1.0:  return f"{p*100:.1f}%"
        if 1.0 < p <= 100.0: return f"{p:.1f}%"
        return f"{p:.3f}"
    except:
        return "-"

def get_ticker_yf_from_pos(p: dict) -> str:
    return p.get("ticker_yf") or ensure_mx(p.get("ticker",""))

# ========= MULTI-FUENTE PRECIOS con DIAGNÓSTICO =========
INVESTING_MAP: Dict[str, Tuple[str, str]] = {
    # "CEMEXCPO.MX": ("Cemex CPO", "mexico"),
}

def investing_symbol_from_yf(ticker_mx: str) -> Tuple[str, str]:
    if ticker_mx in INVESTING_MAP:
        return INVESTING_MAP[ticker_mx]
    base = ticker_mx.replace(".MX", "")
    return base, "mexico"

def yf_last_close_dict(tickers_mx: List[str]) -> Tuple[Dict[str, Tuple[float, str]], Dict[str, str]]:
    out: Dict[str, Tuple[float, str]] = {}
    diag: Dict[str, str] = {}
    if not tickers_mx: return out, diag
    try:
        data = yf.download(
            tickers=tickers_mx, period="5d", interval="1d",
            progress=False, group_by="ticker", threads=True, auto_adjust=False
        )
        for t in tickers_mx:
            try:
                s = data["Close"] if len(tickers_mx) == 1 else data[t]["Close"]
                val = s.dropna().iloc[-1]
                out[t] = (float(val), "yf")
                diag[t] = "yf_ok"
            except Exception:
                diag[t] = "yf_no_close"
    except Exception as e:
        for t in tickers_mx: diag[t] = f"yf_error:{e}"
    return out, diag

def investing_last_close(ticker_mx: str) -> Tuple[float, bool, str]:
    try:
        import investpy
        try:
            from investpy.search import search_quotes
        except Exception:
            search_quotes = None
    except Exception as e:
        return (0.0, False, f"inv_missing_lib:{e}")

    sym, country = investing_symbol_from_yf(ticker_mx)
    # search_quotes
    if search_quotes is not None:
        try:
            q = search_quotes(text=sym, products=["stocks"], countries=[country], n_results=5)
            if not isinstance(q, list): q = [q]
            for qi in q:
                try:
                    df = qi.retrieve_historical_data(
                        from_date="01/01/2020", to_date=datetime.now().strftime("%d/%m/%Y")
                    )
                    if df is not None and not df.empty:
                        return (float(df["Close"].dropna().iloc[-1]), True, "inv_search_ok")
                except Exception:
                    continue
        except Exception as e:
            pass
    # fallback directo
    try:
        import investpy
        df = investpy.get_stock_historical_data(
            stock=sym, country=country,
            from_date="01/01/2020", to_date=datetime.now().strftime("%d/%m/%Y")
        )
        if df is not None and not df.empty:
            return (float(df["Close"].dropna().iloc[-1]), True, "inv_direct_ok")
        return (0.0, False, "inv_no_data")
    except Exception as e:
        return (0.0, False, f"inv_error:{e}")

def investing_last_close_dict(tickers_mx: List[str]) -> Tuple[Dict[str, Tuple[float, str]], Dict[str, str]]:
    out: Dict[str, Tuple[float, str]] = {}
    diag: Dict[str, str] = {}
    for t in tickers_mx:
        val, ok, d = investing_last_close(t)
        if ok:
            out[t] = (val, "investing")
            diag[t] = d
        else:
            diag[t] = d
    return out, diag

def fetch_prices_multi(tickers_mx: List[str], prefer: str = "yf") -> Tuple[Dict[str, Tuple[float, str]], Dict[str, str]]:
    """
    Devuelve:
      - precios: { ticker_mx: (precio, fuente) }
      - diag:    { ticker_mx: "cadena explicando qué se intentó" }
    """
    res: Dict[str, Tuple[float, str]] = {}
    diag: Dict[str, str] = {}
    if not tickers_mx: return res, diag

    if prefer == "yf":
        a, da = yf_last_close_dict(tickers_mx); res.update(a); diag.update(da)
        need = [t for t in tickers_mx if t not in res]
        b, db = investing_last_close_dict(need); res.update(b)
        for k,v in db.items(): diag[k] = (diag.get(k,"") + "|" + v).strip("|")
    else:
        a, da = investing_last_close_dict(tickers_mx); res.update(a); diag.update(da)
        need = [t for t in tickers_mx if t not in res]
        b, db = yf_last_close_dict(need); res.update(b)
        for k,v in db.items(): diag[k] = (diag.get(k,"") + "|" + v).strip("|")

    # ffill
    missing = [t for t in tickers_mx if t not in res]
    if missing:
        try:
            data = yf.download(
                tickers=missing, period="30d", interval="1d",
                progress=False, group_by="ticker", threads=True, auto_adjust=False
            )
            for t in missing:
                try:
                    s = data["Close"].ffill() if len(missing) == 1 else data[t]["Close"].ffill()
                    if not s.dropna().empty:
                        res[t] = (float(s.dropna().iloc[-1]), "yf-ffill")
                        diag[t] = (diag.get(t,"") + "|ffill_ok").strip("|")
                    else:
                        diag[t] = (diag.get(t,"") + "|ffill_empty").strip("|")
                except Exception as e:
                    diag[t] = (diag.get(t,"") + f"|ffill_error:{e}").strip("|")
        except Exception as e:
            for t in missing:
                diag[t] = (diag.get(t,"") + f"|ffill_batch_error:{e}").strip("|")
    return res, diag

# ========= Excel =========
def ensure_excel_with_headers(path: str, sheet_name: str, headers: list[str]):
    if not os.path.exists(path):
        wb = Workbook(); ws = wb.active; ws.title = sheet_name; ws.append(headers); wb.save(path); return
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name); ws.append(headers); wb.save(path); return
    ws = wb[sheet_name]
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(headers); wb.save(path)

def append_excel_row(path: str, sheet_name: str, row_values: list):
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames: wb.create_sheet(title=sheet_name)
    wb[sheet_name].append(row_values); wb.save(path)

def now_ts_fields():
    dt = datetime.now()
    return dt.isoformat(), dt.date().isoformat(), dt.strftime("%H:%M:%S"), f"{dt.minute:02d}"

def ensure_skips_sheet(excel_path: str):
    ensure_excel_with_headers(
        excel_path, "skips",
        [
            "ts","date","time","minute","month","ticker","ticker_yf","side",
            "reason_code","reason_msg","price_diag",
            "rr","min_rr","atr_used","sl_pct","tp_pct",
            "entry_price","price_source","per_trade_cash",
            "budget_used","budget_total","max_open","active_count"
        ]
    )

def log_skip_excel(excel_path: str, month: str, ticker: str, ticker_yf: str, side: str,
                   reason_code: str, reason_msg: str, price_diag: str = "",
                   rr=None, min_rr=None, atr_used=None, sl_pct=None, tp_pct=None,
                   entry_price=None, price_source=None,
                   per_trade_cash=None, budget_used=None, budget_total=None,
                   max_open=None, active_count=None):
    ts, d, t, m = now_ts_fields()
    append_excel_row(
        excel_path, "skips",
        [ts, d, t, m, month, ticker, ticker_yf, side,
         reason_code, reason_msg, price_diag,
         rr, min_rr, atr_used, sl_pct, tp_pct,
         entry_price, price_source, per_trade_cash,
         budget_used, budget_total, max_open, active_count]
    )

# ========= Política =========
def load_policy(month: str, base_dir: str):
    wfjson = os.path.join(base_dir, "reports", "forecast", month, "policy_wfbox.json")
    if os.path.exists(wfjson):
        p = read_json(wfjson)
        return {
            "tp_pct": safe_float(p.get("tp_pct")),
            "sl_pct": safe_float(p.get("sl_pct")),
            "horizon_days": int(p.get("horizon_days", 5)),
            "per_trade_cash": safe_float(p.get("per_trade_cash", 2000)),
            "max_open": int(p.get("max_open", 5)),
            "budget": safe_float(p.get("budget", 10000)),
        }
    wf_csv = os.path.join(base_dir, "wf_box", "reports", "forecast", "policy_selected_walkforward.csv")
    if not os.path.exists(wf_csv):
        raise FileNotFoundError("No hay política: faltan policy_wfbox.json y wf_box/.../policy_selected_walkforward.csv")
    df = pd.read_csv(wf_csv)
    if "month" not in df.columns: raise ValueError("El CSV de walk-forward no tiene columna 'month'")
    row = df.loc[df["month"] == month]
    if row.empty: raise ValueError(f"No encontré política para el mes {month} en {wf_csv}")
    r = row.iloc[0].to_dict()
    return {
        "tp_pct": safe_float(r.get("tp_pct")),
        "sl_pct": safe_float(r.get("sl_pct")),
        "horizon_days": int(safe_float(r.get("horizon_days"), 5)),
        "per_trade_cash": safe_float(r.get("per_trade_cash", 2000)),
        "max_open": int(safe_float(r.get("max_open"), 5)),
        "budget": safe_float(r.get("budget", 10000)),
    }

# ========= Forecast =========
def load_forecast(month: str, base_dir: str) -> pd.DataFrame:
    path = os.path.join(base_dir, "reports", "forecast", month, f"forecast_{month}_with_gate.csv")
    if not os.path.exists(path):
        raise FileNotFoundError(f"No existe {path}.")
    df = pd.read_csv(path)
    if "ticker" not in df.columns:
        if "symbol" in df.columns: df = df.rename(columns={"symbol":"ticker"})
        else: raise ValueError("No 'ticker'/'symbol'.")
    date_col = None
    for c in ["date","dt","signal_date","session_date"]:
        if c in df.columns: date_col = c; break
    if date_col is None: raise ValueError("No fecha en forecast.")
    df["date"] = pd.to_datetime(df[date_col])
    if "side" not in df.columns:
        for c in ["signal","action"]:
            if c in df.columns: df["side"] = df[c].str.upper()
    if "side" not in df.columns: raise ValueError("No side.")
    score_col = None
    for c in ["abs_y","prob","score","abs_score","yhat_prob"]:
        if c in df.columns: score_col = c; break
    if score_col is None: df["__score__"]=1.0; score_col="__score__"
    df["score"] = pd.to_numeric(df[score_col], errors="coerce").fillna(0.0)
    return df[["ticker","date","side","score"]].copy()

def pick_today_signals(df: pd.DataFrame, today: date, max_candidates: int = 10, only_buy: bool = True) -> pd.DataFrame:
    chosen_day = latest_business_day_in(df["date"], today)
    if chosen_day is None: return df.iloc[0:0].copy()
    day = df.loc[df["date"].dt.date == chosen_day.date()].copy()
    if day.empty: return day
    if only_buy: day = day[day["side"].str.upper()=="BUY"].copy()
    day = day.sort_values(["ticker","score"], ascending=[True, False]).groupby("ticker", as_index=False).first()
    return day.sort_values("score", ascending=False).head(max_candidates)

# ========= ATR & Targets =========
def compute_atr_14(ticker_mx: str, lookback_days: int = 60) -> float | None:
    try:
        df = yf.download(tickers=ticker_mx, period=f"{lookback_days}d", interval="1d",
                         progress=False, group_by="ticker", threads=True, auto_adjust=False)
        if df is None or df.empty: return None
        if isinstance(df.columns, pd.MultiIndex):
            try: df = df[ticker_mx]
            except Exception: pass
        high = df["High"].astype(float); low = df["Low"].astype(float); close = df["Close"].astype(float)
        prev_close = close.shift(1)
        tr = pd.concat([(high - low).abs(), (high - prev_close).abs(), (low - prev_close).abs()], axis=1).max(axis=1)
        atr = tr.ewm(alpha=1/14, adjust=False).mean()
        return float(atr.dropna().iloc[-1]) if not atr.dropna().empty else None
    except Exception as e:
        print(f"[ATR] Error {ticker_mx}: {e}"); return None

def compute_tp_sl_buy_hybrid(entry: float, atr: float | None, atr_mult: float,
                             fallback_sl_pct: float, min_rr: float, tp_pct_policy: float | None,
                             tp_ceiling_atr_mult: float | None = None):
    fallback_sl_abs = entry * float(fallback_sl_pct or 0.0)
    atr_sl_abs = (atr or 0.0) * float(atr_mult or 0.0)
    risk_per_share = max(fallback_sl_abs, atr_sl_abs)
    sl = max(0.01, entry - risk_per_share)

    tp_rr = entry + float(min_rr) * (entry - sl)
    tp_pct_abs = entry * (1.0 + float(tp_pct_policy or 0.0))
    tp = max(tp_rr, tp_pct_abs)

    if tp_ceiling_atr_mult is not None and atr is not None:
        tp_ceiling = entry + float(tp_ceiling_atr_mult) * atr
        tp = min(tp, tp_ceiling)

    risk_per_share = entry - sl
    rr = (tp - entry) / max(1e-9, risk_per_share)
    tp_pct_vs_entry = (tp - entry) / entry
    sl_pct_vs_entry = (entry - sl) / entry
    return sl, tp, risk_per_share, rr, tp_pct_vs_entry, sl_pct_vs_entry

# ========= Broker (solo BUY) =========
class PaperBroker:
    def __init__(self, per_trade_cash: float, max_open: int, budget: float,
                 state_file="open_positions.json", excel_path="trading_log.xlsx",
                 unstuck_max_misses: int = 6, unstuck_action: str = "close"):
        """
        state_file: NUEVO archivo de posiciones (reemplaza active_positions.json). Migra si existe.
        unstuck_action: 'close' (cierre técnico al entry), 'mark' (solo marcar), 'alert' (solo Telegram).
        """
        self.per_trade_cash = per_trade_cash
        self.max_open = max_open
        self.budget = budget
        self.state_file = state_file
        self.excel_path = excel_path
        self.unstuck_max_misses = int(unstuck_max_misses)
        self.unstuck_action = unstuck_action
        self._load()
        # Excel headers
        ensure_excel_with_headers(
            self.excel_path,"opens",
            ["opened_at","opened_date","opened_time","opened_minute",
             "id","ticker","ticker_yf","side",
             "entry_price","price_source_open","qty","investment",
             "sl","sl_pct","tp","tp_pct","rr",
             "prob","atr_used","cash_available_post_open","month"]
        )
        ensure_excel_with_headers(
            self.excel_path,"closes",
            ["closed_at","closed_date","closed_time","closed_minute",
             "id","ticker","ticker_yf","side",
             "entry_price","exit_price","price_source_close","qty","investment",
             "pnl","pnl_pct","result","held_days",
             "sl","tp","rr","prob","atr_used","month"]
        )
        ensure_skips_sheet(self.excel_path)

    def _load(self):
        # migración desde active_positions.json si existe
        legacy = "active_positions.json"
        if os.path.exists(self.state_file):
            with open(self.state_file,"r",encoding="utf-8") as f: self.positions = json.load(f)
        elif os.path.exists(legacy):
            with open(legacy,"r",encoding="utf-8") as f: self.positions = json.load(f)
            with open(self.state_file,"w",encoding="utf-8") as f: json.dump(self.positions,f,indent=2,ensure_ascii=False)
        else:
            self.positions = {}
        # normaliza
        changed = False
        for pid,p in list(self.positions.items()):
            if "ticker_yf" not in p and "ticker" in p: p["ticker_yf"]=ensure_mx(p["ticker"]); changed=True
            if "rr" not in p: p["rr"]=None; changed=True
            if "prob" not in p: p["prob"]=None; changed=True
            if "atr_used" not in p: p["atr_used"]=None; changed=True
            if "month" not in p: p["month"]=""; changed=True
            if "miss_no_price" not in p: p["miss_no_price"]=0; changed=True
        if changed: self._save()

    def _save(self):
        with open(self.state_file,"w",encoding="utf-8") as f:
            json.dump(self.positions,f,indent=2,ensure_ascii=False)

    def active_count(self):
        return sum(1 for p in self.positions.values() if p["status"]=="ACTIVE")

    def used_budget(self):
        return sum(p["investment"] for p in self.positions.values() if p["status"]=="ACTIVE")

    @staticmethod
    def _split_dt_fields(dt_iso: str):
        dt = datetime.fromisoformat(dt_iso)
        return dt.date().isoformat(), dt.strftime("%H:%M:%S"), f"{dt.minute:02d}"

    def log_open_excel(self, p: dict):
        d,t,mn = self._split_dt_fields(p["opened_at"])
        append_excel_row(self.excel_path,"opens",
            [p["opened_at"],d,t,mn,p["id"],p["ticker"],get_ticker_yf_from_pos(p),p["side"],
             p["entry_price"],p.get("price_source_open",""),p["qty"],p["investment"],
             p["sl"],p["sl_pct"],p["tp"],p["tp_pct"],p["rr"],
             p.get("prob",None),p.get("atr_used",None),p.get("cash_available_post_open",None),p.get("month","")]
        )

    def log_close_excel(self, p: dict, pnl: float, pnl_pct: float, held_days: int, price_source_close: str = ""):
        d,t,mn = self._split_dt_fields(p["exit_at"])
        append_excel_row(self.excel_path,"closes",
            [p["exit_at"],d,t,mn,p["id"],p["ticker"],get_ticker_yf_from_pos(p),p["side"],
             p["entry_price"],p["exit_price"],price_source_close,p["qty"],p["investment"],
             pnl,pnl_pct,p.get("result",""),held_days,
             p["sl"],p["tp"],p.get("rr",None),p.get("prob",None),p.get("atr_used",None),p.get("month","")]
        )

    # Apertura
    def try_open_buy(self, ticker: str, price: float, price_source_open: str,
                     sl: float, tp: float, horizon_days: int,
                     per_trade_cash: float, prob, rr: float, month: str,
                     sl_pct: float, tp_pct: float, atr_used: float, ticker_yf: str):
        if self.active_count() >= self.max_open: return None, "MAX_OPEN_REACHED"
        if self.used_budget() + per_trade_cash > self.budget: return None, "BUDGET_EXCEEDED"
        qty = max(1, int(per_trade_cash // price)); invest = qty * price
        pid = f"{ticker}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.positions[pid] = {
            "id":pid,"ticker":ticker,"ticker_yf":ticker_yf,"side":"BUY",
            "entry_price":price,"price_source_open":price_source_open,
            "tp":tp,"sl":sl,"sl_pct":sl_pct,"tp_pct":tp_pct,
            "qty":qty,"investment":invest,
            "opened_at":datetime.now().isoformat(),"max_holding_days":int(horizon_days),
            "status":"ACTIVE","prob":prob,"rr":rr,"atr_used":atr_used,"month":month,
            "miss_no_price":0
        }
        self._save()
        self.positions[pid]["cash_available_post_open"] = self.budget - self.used_budget()
        self._save()
        self.log_open_excel(self.positions[pid])
        return pid, "OPENED"

    # Cierre
    def _force_close(self, pid: str, last_price: float, reason: str, price_source_close: str = ""):
        p = self.positions[pid]; qty=p["qty"]; entry=p["entry_price"]
        pnl = (last_price - entry) * qty
        p["status"]="CLOSED"; p["exit_price"]=last_price; p["exit_at"]=datetime.now().isoformat()
        p["pnl"]=pnl; p["result"]=reason; self._save()
        held_days=(datetime.fromisoformat(p["exit_at"]) - datetime.fromisoformat(p["opened_at"])).days
        pnl_pct = pnl/max(1e-9,p["investment"])
        self.log_close_excel(p,pnl,pnl_pct,held_days,price_source_close=price_source_close)
        return pnl,pnl_pct,held_days

    def try_close_hits(self, pid: str, last_price: float, price_source_close: str = ""):
        p = self.positions[pid]
        if p["status"]!="ACTIVE": return False,None,0.0,0.0,0
        tp=p["tp"]; sl=p["sl"]; entry=p["entry_price"]
        hit=None
        if last_price >= tp: hit="TP_HIT"
        elif last_price <= sl: hit="SL_HIT"
        if hit:
            p["miss_no_price"]=0
            pnl,pnl_pct,held_days = self._force_close(pid,last_price,hit,price_source_close)
            return True,hit,pnl,pnl_pct,held_days
        opened=datetime.fromisoformat(p["opened_at"])
        if datetime.now() >= opened + timedelta(days=int(p["max_holding_days"])):
            pnl,pnl_pct,held_days = self._force_close(pid,last_price,"TIME_EXIT",price_source_close)
            return True,"TIME_EXIT",pnl,pnl_pct,held_days
        return False,None,0.0,0.0,0

# ========= Cierre masivo =========
def close_all_active_now(broker: PaperBroker, prefer="yf"):
    active = [(pid,p) for pid,p in broker.positions.items() if p["status"]=="ACTIVE"]
    if not active:
        print("No hay posiciones activas para cerrar."); return 0,0.0
    tickers = list({ensure_mx(p["ticker"]) for _,p in active})
    prices, diag = fetch_prices_multi(tickers, prefer=prefer)
    closed_n=0; total_pnl=0.0
    for pid,p in active:
        tmx=ensure_mx(p["ticker"])
        if tmx not in prices:
            print(f"[RESET] Sin precio para {tmx} ({diag.get(tmx,'-')}). Salto.")
            continue
        last_price, source = prices[tmx]
        pnl,pnl_pct,held_days = broker._force_close(pid,last_price,"MANUAL_RESET",price_source_close=source)
        closed_n+=1; total_pnl+=pnl
        send_tg(
            f"⚠️ <b>CIERRE MASIVO</b>\n"
            f"📈 <b>{p['ticker']}</b> ({get_ticker_yf_from_pos(p)}) BUY | Qty: <b>{p['qty']}</b>\n"
            f"📍 MANUAL_RESET | ⏳ {held_days}d | R≈{(p.get('rr') or 0):.2f}x\n"
            f"💰 Entrada: {pretty_cash(p['entry_price'])} → Salida: {pretty_cash(last_price)}\n"
            f"🛰️ Fuente precio: {source}\n"
            f"📊 P&L: <b>{pretty_cash(pnl)}</b> ({pnl_pct*100:+.1f}%)"
        )
    print(f"[RESET] Cerradas {closed_n}. PnL total ~ {pretty_cash(total_pnl)}"); return closed_n,total_pnl

# ========= Monitor =========
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--month", required=True)
    parser.add_argument("--base-dir", default=".")
    parser.add_argument("--interval-min", type=int, default=15)
    parser.add_argument("--max-new", type=int, default=3)

    parser.add_argument("--only-buy", action="store_true", default=True)
    parser.add_argument("--atr-mult", type=float, default=4.0)
    parser.add_argument("--min-rr", type=float, default=4.0)
    parser.add_argument("--fallback-sl-pct", type=float, default=0.001)
    parser.add_argument("--tp-ceiling-atr-mult", type=float, default=5.0)
    parser.add_argument("--excel-path", type=str, default="trading_log.xlsx")

    # NUEVO: archivo de posiciones + auto-unstuck
    parser.add_argument("--positions-file", type=str, default="open_positions.json",
                        help="Ruta del archivo JSON de posiciones abiertas.")
    parser.add_argument("--auto-unstuck", action="store_true", default=True,
                        help="Auto-liberar posiciones 'atascadas' sin precio.")
    parser.add_argument("--unstuck-max-misses", type=int, default=6,
                        help="Ciclos sin precio antes de actuar.")
    parser.add_argument("--unstuck-action", choices=["close","mark","alert"], default="close",
                        help="Acción al superar misses: close=cierrar al entry, mark=marcar, alert=solo Telegram.")

    # Reset
    parser.add_argument("--ask-reset", action="store_true", default=True)
    parser.add_argument("--force-reset", action="store_true", default=False)
    parser.add_argument("--no-ask-reset", action="store_true", default=False)

    # Skips
    parser.add_argument("--log-all-candidates", action="store_true", default=False)
    parser.add_argument("--skips-to-telegram", action="store_true", default=False)

    # Fuente preferida
    parser.add_argument("--prefer", choices=["yf","investing"], default="yf")

    args = parser.parse_args()
    month = args.month
    base_dir = os.path.abspath(args.base_dir)

    # Política
    pol = load_policy(month, base_dir)
    tp_pct_base = float(pol["tp_pct"] or 0.0)
    sl_pct_fallback = float(pol["sl_pct"] or args.fallback_sl_pct)
    horizon_days = int(pol["horizon_days"])
    per_trade_cash = float(pol["per_trade_cash"])
    max_open = int(pol["max_open"])
    budget = float(pol["budget"])

    print("=== BMV Monitor v5 (solo BUY | ATR SL | TP híbrido | Excel + skips + auto-unstuck) ===")
    print(f"Mes {month} | H={horizon_days}d | cash/trade={pretty_cash(per_trade_cash)} | max_open={max_open} | budget={pretty_cash(budget)}")
    print(f"SL: ATR*{args.atr_mult} (respaldo {sl_pct_fallback*100:.2f}%) | Rmin {args.min_rr}x | tp_pct_política={tp_pct_base*100:.1f}% | techo TP ATR*{args.tp_ceiling_atr_mult}")
    print(f"Excel: {args.excel_path} | Posiciones: {args.positions_file} | prefer: {args.prefer}")
    print(f"Auto-unstuck: {args.auto_unstuck} (max_misses={args.unstuck_max_misses}, action={args.unstuck_action})")

    send_tg(
        f"🚀 <b>MONITOR BMV v5</b>\n"
        f"Mes {month} | H {horizon_days}d | 💵 {pretty_cash(per_trade_cash)} por trade | max_open={max_open} | budget={pretty_cash(budget)}\n"
        f"SL ATR*{args.atr_mult} · RR≥{args.min_rr}x · TP híbrido (min política {tp_pct_base*100:.1f}%)"
    )

    broker = PaperBroker(
        per_trade_cash=per_trade_cash, max_open=max_open, budget=budget,
        state_file=args.positions_file, excel_path=args.excel_path,
        unstuck_max_misses=args.unstuck_max_misses, unstuck_action=args.unstuck_action
    )

    # Reset inicio
    if args.force_reset:
        print("[RESET] Forzado."); close_all_active_now(broker, prefer=args.prefer)
    elif not args.no_ask_reset and args.ask_reset and broker.active_count()>0:
        try: resp=input("¿Cerrar TODAS las posiciones activas ahora para reiniciar? [s/N]: ").strip().lower()
        except EOFError: resp="n"
        if resp in ("s","si","sí","y","yes"): close_all_active_now(broker, prefer=args.prefer)
        else: print("[RESET] No se cerraron posiciones al inicio.")

    # Forecast
    df_fc = load_forecast(month, base_dir)
    today = datetime.now().date()
    cand = pick_today_signals(df_fc, today, max_candidates=50, only_buy=args.only_buy)
    if cand.empty:
        print("No hay señales BUY para hoy."); send_tg("ℹ️ No hay señales BUY para hoy."); return
    print(f"Candidatas BUY hoy: {len(cand)}"); print(cand.head(10).to_string(index=False))

    while True:
        try:
            # 1) CIERRES
            active = [(pid,p) for pid,p in broker.positions.items() if p["status"]=="ACTIVE"]
            if active:
                print(f"[{datetime.now():%H:%M:%S}] Activas: {len(active)}")
                tickers = list({ensure_mx(p["ticker"]) for _,p in active})
                prices, diag = fetch_prices_multi(tickers, prefer=args.prefer)

                for pid,p in active:
                    tmx = ensure_mx(p["ticker"])
                    if tmx in prices:
                        last_price, source = prices[tmx]
                        p["miss_no_price"]=0; broker._save()
                        closed,reason,pnl,pnl_pct,held_days = broker.try_close_hits(pid,last_price,price_source_close=source)
                        if closed:
                            emoji = "🎉" if reason=="TP_HIT" else ("🛑" if reason=="SL_HIT" else "⏰")
                            send_tg(
                                f"{emoji} <b>POSICIÓN CERRADA</b>\n"
                                f"📈 <b>{p['ticker']}</b> ({get_ticker_yf_from_pos(p)}) BUY | Qty: <b>{p['qty']}</b>\n"
                                f"📍 {reason} | ⏳ {held_days}d | R≈{(p.get('rr') or 0):.2f}x\n"
                                f"💰 Entrada: {pretty_cash(p['entry_price'])} → Salida: {pretty_cash(broker.positions[pid]['exit_price'])}\n"
                                f"🛰️ Fuente: {source}\n"
                                f"📊 P&L: <b>{pretty_cash(pnl)}</b> ({pnl_pct*100:+.1f}%)"
                            )
                    else:
                        # no price → aumenta miss y tomar acción si excede umbral
                        p["miss_no_price"] = int(p.get("miss_no_price",0)) + 1; broker._save()
                        log_skip_excel(
                            broker.excel_path, month, p["ticker"], tmx, "BUY",
                            reason_code="no_price_close",
                            reason_msg="No hay precio para evaluar cierre",
                            price_diag=diag.get(tmx,""),
                            rr=p.get("rr"), min_rr=None, atr_used=p.get("atr_used"),
                            sl_pct=p.get("sl_pct"), tp_pct=p.get("tp_pct"),
                            entry_price=p.get("entry_price"), price_source=None,
                            per_trade_cash=broker.per_trade_cash,
                            budget_used=broker.used_budget(), budget_total=broker.budget,
                            max_open=max_open, active_count=broker.active_count()
                        )
                        # AUTO-UNSTUCK
                        if args.auto_unstuck and p["miss_no_price"] >= broker.unstuck_max_misses:
                            action = broker.unstuck_action
                            if action == "close":
                                # cierre técnico al ENTRY
                                entry = p["entry_price"]
                                pnl,pnl_pct,held_days = broker._force_close(pid, entry, "AUTO_UNSTUCK_FLAT", price_source_close="unstuck-flat")
                                send_tg(
                                    f"🧹 <b>AUTO-UNSTUCK</b> (close)\n"
                                    f"📈 {p['ticker']} ({tmx}) | misses={p['miss_no_price']}\n"
                                    f"Se cierra al precio de entrada por falta de datos.\n"
                                    f"📊 P&L: {pretty_cash(pnl)} ({pnl_pct*100:+.1f}%)"
                                )
                            elif action == "mark":
                                p["status"]="STUCK"; broker._save()
                                send_tg(f"🧹 <b>AUTO-UNSTUCK</b> (mark) {p['ticker']} ({tmx}) marcado como STUCK. misses={p['miss_no_price']}")
                            elif action == "alert":
                                send_tg(f"🧹 <b>AUTO-UNSTUCK</b> (alert) {p['ticker']} ({tmx}) supera misses={p['miss_no_price']} sin precio.")

            # 2) APERTURAS
            capacity = max(0, max_open - broker.active_count())
            to_open = min(capacity, args.max_new)

            if to_open <= 0 and args.log_all_candidates:
                active_tickers = {p["ticker"] for p in broker.positions.values() if p["status"]=="ACTIVE"}
                pending = cand[~cand["ticker"].isin(active_tickers)]
                for _,row in pending.iterrows():
                    t=row["ticker"]; tmx=ensure_mx(t)
                    log_skip_excel(
                        broker.excel_path, month, t, tmx, "BUY",
                        reason_code="capacity_full",
                        reason_msg=f"Sin capacidad: active={broker.active_count()} / max_open={max_open}",
                        price_diag="", per_trade_cash=broker.per_trade_cash,
                        budget_used=broker.used_budget(), budget_total=broker.budget,
                        max_open=max_open, active_count=broker.active_count()
                    )

            if to_open > 0:
                active_tickers = {p["ticker"] for p in broker.positions.values() if p["status"]=="ACTIVE"}
                open_candidates = cand[~cand["ticker"].isin(active_tickers)].head(to_open)

                if args.log_all_candidates:
                    skipped_active = cand[cand["ticker"].isin(active_tickers)]
                    for _,row in skipped_active.iterrows():
                        t=row["ticker"]; tmx=ensure_mx(t)
                        log_skip_excel(
                            broker.excel_path, month, t, tmx, "BUY",
                            reason_code="already_active",
                            reason_msg="Ticker ya tiene posición activa",
                            price_diag="", per_trade_cash=broker.per_trade_cash,
                            budget_used=broker.used_budget(), budget_total=broker.budget,
                            max_open=max_open, active_count=broker.active_count()
                        )

                if not open_candidates.empty:
                    tickers_new = [ensure_mx(t) for t in open_candidates["ticker"].tolist()]
                    price_dict, diag_open = fetch_prices_multi(tickers_new, prefer=args.prefer)

                    opened = []
                    skipped_cycle = []

                    for _,row in open_candidates.iterrows():
                        t=row["ticker"]; tmx=ensure_mx(t)
                        if tmx not in price_dict:
                            if args.log_all_candidates:
                                log_skip_excel(
                                    broker.excel_path, month, t, tmx, "BUY",
                                    reason_code="no_price",
                                    reason_msg="No hay precio (YF/INV/ffill)",
                                    price_diag=diag_open.get(tmx,""),
                                    per_trade_cash=broker.per_trade_cash,
                                    budget_used=broker.used_budget(), budget_total=broker.budget,
                                    max_open=max_open, active_count=broker.active_count()
                                )
                            skipped_cycle.append(("no_price"))
                            continue

                        price, price_source = price_dict[tmx]
                        atr_val = compute_atr_14(tmx, lookback_days=60)
                        sl,tp,risk_ps,rr,tp_pct_vs_entry,sl_pct_vs_entry = compute_tp_sl_buy_hybrid(
                            entry=price, atr=atr_val, atr_mult=args.atr_mult,
                            fallback_sl_pct=sl_pct_fallback, min_rr=args.min_rr,
                            tp_pct_policy=tp_pct_base, tp_ceiling_atr_mult=args.tp_ceiling_atr_mult
                        )
                        if rr < args.min_rr:
                            if args.log_all_candidates:
                                log_skip_excel(
                                    broker.excel_path, month, t, tmx, "BUY",
                                    reason_code="rr_too_low",
                                    reason_msg=f"R {rr:.2f} < min_rr {args.min_rr:.2f}",
                                    price_diag=diag_open.get(tmx,""),
                                    rr=rr, min_rr=args.min_rr, atr_used=atr_val,
                                    sl_pct=sl_pct_vs_entry, tp_pct=tp_pct_vs_entry,
                                    entry_price=price, price_source=price_source,
                                    per_trade_cash=broker.per_trade_cash,
                                    budget_used=broker.used_budget(), budget_total=broker.budget,
                                    max_open=max_open, active_count=broker.active_count()
                                )
                            skipped_cycle.append(("rr_too_low"))
                            continue

                        pid,status = broker.try_open_buy(
                            ticker=t, price=price, price_source_open=price_source,
                            sl=sl, tp=tp, horizon_days=horizon_days, per_trade_cash=per_trade_cash,
                            prob=float(row.get("score",0.0)), rr=rr, month=month,
                            sl_pct=sl_pct_vs_entry, tp_pct=tp_pct_vs_entry,
                            atr_used=float(atr_val) if atr_val is not None else None, ticker_yf=tmx
                        )
                        if status == "OPENED":
                            opened.append(tmx)
                            p=broker.positions[pid]; qty=p["qty"]; inv=p["investment"]; pnl_target=(tp-price)*qty
                            send_tg(
                                f"🚀 <b>NUEVA POSICIÓN</b>\n"
                                f"📈 <b>{t}</b> ({tmx}) BUY | Fuente: {price_source}\n"
                                f"💰 Entrada: {pretty_cash(price)} | Qty: <b>{qty}</b> | Inversión: {pretty_cash(inv)}\n"
                                f"🛡️ SL: {pretty_cash(sl)} ({sl_pct_vs_entry*100:.2f}%)  🎯 TP: {pretty_cash(tp)} ({tp_pct_vs_entry*100:.2f}%)\n"
                                f"🔢 R≈{rr:.2f}x  🔮 Prob: {prob_to_pct(p.get('prob'))}\n"
                                f"📏 ATR: {pretty_cash(p.get('atr_used') or 0.0)}  💼 Cash post: {pretty_cash(p.get('cash_available_post_open',0.0))}\n"
                                f"📊 Ganancia estimada a TP: <b>{pretty_cash(pnl_target)}</b>  ⏳ H: {horizon_days}d"
                            )
                        else:
                            reason = "max_open_reached" if status=="MAX_OPEN_REACHED" else ("budget_exceeded" if status=="BUDGET_EXCEEDED" else "other")
                            if args.log_all_candidates:
                                log_skip_excel(
                                    broker.excel_path, month, t, tmx, "BUY",
                                    reason_code=reason, reason_msg=f"status={status}",
                                    price_diag=diag_open.get(tmx,""),
                                    rr=rr, min_rr=args.min_rr, atr_used=atr_val,
                                    sl_pct=sl_pct_vs_entry, tp_pct=tp_pct_vs_entry,
                                    entry_price=price, price_source=price_source,
                                    per_trade_cash=broker.per_trade_cash,
                                    budget_used=broker.used_budget(), budget_total=broker.budget,
                                    max_open=max_open, active_count=broker.active_count()
                                )
                            skipped_cycle.append((reason))

                    if args.skips_to_telegram and skipped_cycle:
                        cnt = Counter(skipped_cycle)
                        detalle = "\n".join([f"• {k}: {v}" for k,v in cnt.items()])
                        send_tg(f"ℹ️ <b>Descartes apertura (ciclo)</b>\n{detalle}")

            # 3) Espera
            wake = (datetime.now()+timedelta(minutes=args.interval_min)).strftime("%H:%M")
            print(f"Duerme {args.interval_min} min - {wake}")
            time.sleep(args.interval_min*60)

        except KeyboardInterrupt:
            print("Detenido por usuario."); send_tg("🛑 Monitor BMV detenido por el usuario."); break
        except Exception as e:
            print(f"Error ciclo: {e}"); time.sleep(10)

if __name__ == "__main__":
    main()
