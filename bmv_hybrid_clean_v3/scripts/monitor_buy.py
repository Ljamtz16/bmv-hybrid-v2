# monitor_bmv_buy_atr_rr_hybrid_excel_v4.py
import os, json, time, argparse
from datetime import datetime, timedelta, date
from typing import Dict, Tuple, List
import pandas as pd
import numpy as np
import yfinance as yf
import requests
from dotenv import load_dotenv, find_dotenv
from openpyxl import Workbook, load_workbook
from collections import Counter, defaultdict

# ========= Telegram =========
load_dotenv()
env_path = find_dotenv(usecwd=True)
loaded = load_dotenv(env_path, override=True)
print("ENV loaded:", loaded, "| path:", env_path or "<none>")

TOKEN = (os.getenv("TELEGRAM_BOT_TOKEN") or "").strip()
CHAT_ID = (os.getenv("TELEGRAM_CHAT_ID") or "").strip()

def send_tg(msg: str) -> bool:
    if not TOKEN or not CHAT_ID:
        print(f"[TG] Faltan credenciales: TOKEN={'OK' if TOKEN else 'MISSING'} CHAT_ID={CHAT_ID}")
        return False
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{TOKEN}/sendMessage",
            data={"chat_id": CHAT_ID, "text": msg, "parse_mode": "HTML"},
            timeout=15,
        )
        ok = r.status_code == 200 and r.json().get("ok", False)
        if not ok:
            print(f"[TG] Error: HTTP {r.status_code} {r.text}")
        return ok
    except Exception as e:
        print(f"[TG] Excepción: {e}")
        return False

# ========= Utilidades =========
def read_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def safe_float(x, default=None):
    try:
        if x is None: return default
        if isinstance(x, (int, float)): return float(x)
        return float(str(x).replace(",", ""))
    except:
        return default

def ensure_mx(ticker: str) -> str:
    t = ticker.strip()
    return t if t.endswith(".MX") else f"{t}.MX"

def latest_business_day_in(df_dates: pd.Series, until: date) -> pd.Timestamp | None:
    d = pd.to_datetime(df_dates).dt.date
    d = d[d <= until]
    if d.empty: return None
    return pd.Timestamp(max(d))

def pretty_cash(x: float) -> str:
    return f"${x:,.2f}"

def pretty_pct(x: float) -> str:
    return f"{x*100:.1f}%"

def prob_to_pct(prob):
    try:
        p = float(prob)
        if 0.0 <= p <= 1.0:
            return f"{p*100:.1f}%"
        if 1.0 < p <= 100.0:
            return f"{p:.1f}%"
        return f"{p:.3f}"
    except:
        return "-"

def get_ticker_yf_from_pos(p: dict) -> str:
    t = p.get("ticker_yf")
    if t:
        return t
    return ensure_mx(p.get("ticker",""))

# ========= MULTI-FUENTE PRECIOS: YF + Investing =========
# Mapea tickers BMV a nombres de Investing.com cuando no coinciden:
INVESTING_MAP: Dict[str, Tuple[str, str]] = {
    # "CEMEXCPO.MX": ("Cemex CPO", "mexico"),
    # Agrega aquí los que necesites
}

def investing_symbol_from_yf(ticker_mx: str) -> Tuple[str, str]:
    if ticker_mx in INVESTING_MAP:
        return INVESTING_MAP[ticker_mx]
    base = ticker_mx.replace(".MX", "")
    return base, "mexico"

def yf_last_close_dict(tickers_mx: List[str]) -> Dict[str, Tuple[float, str]]:
    out: Dict[str, Tuple[float, str]] = {}
    if not tickers_mx:
        return out
    try:
        data = yf.download(
            tickers=tickers_mx,
            period="5d",
            interval="1d",
            progress=False,
            group_by="ticker",
            threads=True,
            auto_adjust=False
        )
        for t in tickers_mx:
            try:
                if len(tickers_mx) == 1:
                    s = data["Close"]
                else:
                    s = data[t]["Close"]
                val = s.dropna().iloc[-1]
                out[t] = (float(val), "yf")
            except Exception:
                pass
    except Exception:
        pass
    return out

def investing_last_close(ticker_mx: str) -> Tuple[float, bool]:
    try:
        import investpy
        try:
            search_quotes = getattr(__import__('investpy'), 'search_quotes', None)
        except Exception:
            search_quotes = None
    except Exception:
        return (0.0, False)

    try:
        sym, country = investing_symbol_from_yf(ticker_mx)
        if search_quotes is not None:
            try:
                q = search_quotes(text=sym, products=["stocks"], countries=[country], n_results=5)
                if not isinstance(q, list):
                    q = [q]
                for qi in q:
                    try:
                        df = qi.retrieve_historical_data(
                            from_date="01/01/2020",
                            to_date=datetime.now().strftime("%d/%m/%Y")
                        )
                        if df is not None and not df.empty:
                            val = float(df["Close"].dropna().iloc[-1])
                            return (val, True)
                    except Exception:
                        continue
            except Exception:
                pass
        # fallback directo si conoces el símbolo exacto
        try:
            df = investpy.get_stock_historical_data(
                stock=sym, country=country,
                from_date="01/01/2020",
                to_date=datetime.now().strftime("%d/%m/%Y")
            )
            if df is not None and not df.empty:
                val = float(df["Close"].dropna().iloc[-1])
                return (val, True)
        except Exception:
            pass
    except Exception:
        return (0.0, False)
    return (0.0, False)

def investing_last_close_dict(tickers_mx: List[str]) -> Dict[str, Tuple[float, str]]:
    out: Dict[str, Tuple[float, str]] = {}
    for t in tickers_mx:
        val, ok = investing_last_close(t)
        if ok:
            out[t] = (val, "investing")
    return out

def fetch_prices_multi(tickers_mx: List[str], prefer: str = "yf") -> Dict[str, Tuple[float, str]]:
    """
    Devuelve {ticker_mx: (precio, fuente)}. prefer: 'yf' o 'investing'
    Orden: preferida -> alternativa -> ffill (YF).
    """
    res: Dict[str, Tuple[float, str]] = {}
    if not tickers_mx:
        return res

    if prefer == "yf":
        a = yf_last_close_dict(tickers_mx)
        need = [t for t in tickers_mx if t not in a]
        b = investing_last_close_dict(need)
        res.update(a); res.update(b)
    else:
        a = investing_last_close_dict(tickers_mx)
        need = [t for t in tickers_mx if t not in a]
        b = yf_last_close_dict(need)
        res.update(a); res.update(b)

    # ffill como último recurso
    missing = [t for t in tickers_mx if t not in res]
    if missing:
        try:
            data = yf.download(
                tickers=missing, period="30d", interval="1d",
                progress=False, group_by="ticker", threads=True, auto_adjust=False
            )
            for t in missing:
                try:
                    if len(missing) == 1:
                        s = data["Close"].ffill()
                    else:
                        s = data[t]["Close"].ffill()
                    if not s.dropna().empty:
                        res[t] = (float(s.dropna().iloc[-1]), "yf-ffill")
                except Exception:
                    pass
        except Exception:
            pass
    return res

# ========= Excel logging =========
def ensure_excel_with_headers(path: str, sheet_name: str, headers: list[str]):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        wb.save(path)
        return
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        wb.save(path)
        return
    ws = wb[sheet_name]
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(headers)
        wb.save(path)

def append_excel_row(path: str, sheet_name: str, row_values: list):
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    ws = wb[sheet_name]
    ws.append(row_values)
    wb.save(path)

# ========= Helpers timestamp / SKIPS =========
def now_ts_fields():
    dt = datetime.now()
    return dt.isoformat(), dt.date().isoformat(), dt.strftime("%H:%M:%S"), f"{dt.minute:02d}"

def ensure_skips_sheet(excel_path: str):
    ensure_excel_with_headers(
        excel_path,
        "skips",
        [
            "ts","date","time","minute",
            "month","ticker","ticker_yf","side",
            "reason_code","reason_msg",
            "rr","min_rr","atr_used","sl_pct","tp_pct",
            "entry_price","price_source",
            "per_trade_cash","budget_used","budget_total","max_open","active_count"
        ]
    )

def log_skip_excel(excel_path: str, month: str, ticker: str, ticker_yf: str, side: str,
                   reason_code: str, reason_msg: str,
                   rr=None, min_rr=None, atr_used=None, sl_pct=None, tp_pct=None,
                   entry_price=None, price_source=None,
                   per_trade_cash=None, budget_used=None, budget_total=None,
                   max_open=None, active_count=None):
    ts, d, t, m = now_ts_fields()
    append_excel_row(
        excel_path, "skips",
        [
            ts, d, t, m,
            month, ticker, ticker_yf, side,
            reason_code, reason_msg,
            rr, min_rr, atr_used, sl_pct, tp_pct,
            entry_price, price_source,
            per_trade_cash, budget_used, budget_total, max_open, active_count
        ]
    )

# ========= Política =========
def load_policy(month: str, base_dir: str):
    wfjson = os.path.join(base_dir, "reports", "forecast", month, "policy_wfbox.json")
    if os.path.exists(wfjson):
        p = read_json(wfjson)
        return {
            "tp_pct": safe_float(p.get("tp_pct")),
            "sl_pct": safe_float(p.get("sl_pct")),
            "horizon_days": int(p.get("horizon_days", 5)),
            "per_trade_cash": safe_float(p.get("per_trade_cash", 2000)),
            "max_open": int(p.get("max_open", 5)),
            "budget": safe_float(p.get("budget", 10000)),
        }
    wf_csv = os.path.join(base_dir, "wf_box", "reports", "forecast", "policy_selected_walkforward.csv")
    if not os.path.exists(wf_csv):
        raise FileNotFoundError("No hay política: faltan policy_wfbox.json y wf_box/.../policy_selected_walkforward.csv")
    df = pd.read_csv(wf_csv)
    if "month" not in df.columns:
        raise ValueError("El CSV de walk-forward no tiene columna 'month'")
    row = df.loc[df["month"] == month]
    if row.empty:
        raise ValueError(f"No encontré política para el mes {month} en {wf_csv}")
    r = row.iloc[0].to_dict()
    return {
        "tp_pct": safe_float(r.get("tp_pct")),
        "sl_pct": safe_float(r.get("sl_pct")),
        "horizon_days": int(safe_float(r.get("horizon_days"), 5)),
        "per_trade_cash": safe_float(r.get("per_trade_cash", 2000)),
        "max_open": int(safe_float(r.get("max_open"), 5)),
        "budget": safe_float(r.get("budget", 10000)),
    }

# ========= Forecast / Señales =========
def load_forecast(month: str, base_dir: str) -> pd.DataFrame:
    fc_path = os.path.join(base_dir, "reports", "forecast", month, f"forecast_{month}_with_gate.csv")
    if not os.path.exists(fc_path):
        raise FileNotFoundError(f"No existe {fc_path}. Genera el forecast primero (scripts/12_forecast_and_validate.py).")
    df = pd.read_csv(fc_path)

    if "ticker" not in df.columns:
        if "symbol" in df.columns: df = df.rename(columns={"symbol":"ticker"})
        else: raise ValueError("El forecast no tiene columna 'ticker' (ni 'symbol').")
    date_col = None
    for c in ["date", "dt", "signal_date", "session_date"]:
        if c in df.columns:
            date_col = c; break
    if date_col is None: raise ValueError("No encontré columna de fecha (date/dt/signal_date).")
    df["date"] = pd.to_datetime(df[date_col])

    if "side" not in df.columns:
        for c in ["signal", "action"]:
            if c in df.columns:
                df["side"] = df[c].str.upper()
                break
        if "side" not in df.columns:
            raise ValueError("No encontré columna 'side'/'signal'/'action' (BUY/SELL).")

    score_col = None
    for c in ["abs_y", "prob", "score", "abs_score", "yhat_prob"]:
        if c in df.columns:
            score_col = c; break
    if score_col is None:
        df["__score__"] = 1.0
        score_col = "__score__"

    df["score"] = pd.to_numeric(df[score_col], errors="coerce").fillna(0.0)
    return df[["ticker","date","side","score"]].copy()

def pick_today_signals(df: pd.DataFrame, today: date, max_candidates: int = 10, only_buy: bool = True) -> pd.DataFrame:
    chosen_day = latest_business_day_in(df["date"], today)
    if chosen_day is None:
        return df.iloc[0:0].copy()
    day_df = df.loc[df["date"].dt.date == chosen_day.date()].copy()
    if day_df.empty:
        return day_df
    if only_buy:
        day_df = day_df[day_df["side"].str.upper() == "BUY"].copy()
    day_df = day_df.sort_values(["ticker","score"], ascending=[True, False])
    day_df = day_df.groupby("ticker", as_index=False).first()
    day_df = day_df.sort_values("score", ascending=False).head(max_candidates)
    return day_df

# ========= ATR & Targets =========
def compute_atr_14(ticker_mx: str, lookback_days: int = 60) -> float | None:
    try:
        df = yf.download(tickers=ticker_mx, period=f"{lookback_days}d", interval="1d",
                         progress=False, group_by="ticker", threads=True, auto_adjust=False)
        if df is None or df.empty:
            return None
        if isinstance(df.columns, pd.MultiIndex):
            try:
                df = df[ticker_mx]
            except Exception:
                pass
        high = df["High"].astype(float)
        low  = df["Low"].astype(float)
        close = df["Close"].astype(float)

        prev_close = close.shift(1)
        tr1 = (high - low).abs()
        tr2 = (high - prev_close).abs()
        tr3 = (low - prev_close).abs()
        tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
        atr = tr.ewm(alpha=1/14, adjust=False).mean()
        val = float(atr.dropna().iloc[-1]) if not atr.dropna().empty else None
        return val
    except Exception as e:
        print(f"[ATR] Error {ticker_mx}: {e}")
        return None

def compute_tp_sl_buy_hybrid(entry: float, atr: float | None, atr_mult: float,
                             fallback_sl_pct: float, min_rr: float, tp_pct_policy: float | None,
                             tp_ceiling_atr_mult: float | None = None):
    fallback_sl_abs = entry * float(fallback_sl_pct or 0.0)
    atr_sl_abs = (atr or 0.0) * float(atr_mult or 0.0)
    risk_per_share = max(fallback_sl_abs, atr_sl_abs)
    sl = max(0.01, entry - risk_per_share)

    tp_rr = entry + float(min_rr) * (entry - sl)
    tp_pct_abs = entry * (1.0 + float(tp_pct_policy or 0.0))
    tp = max(tp_rr, tp_pct_abs)

    if tp_ceiling_atr_mult is not None and atr is not None:
        tp_ceiling = entry + float(tp_ceiling_atr_mult) * atr
        tp = min(tp, tp_ceiling)

    risk_per_share = entry - sl
    rr = (tp - entry) / max(1e-9, risk_per_share)
    tp_pct_vs_entry = (tp - entry) / entry
    sl_pct_vs_entry = (entry - sl) / entry
    return sl, tp, risk_per_share, rr, tp_pct_vs_entry, sl_pct_vs_entry

# ========= Broker (solo BUY) =========
class PaperBroker:
    def __init__(self, per_trade_cash: float, max_open: int, budget: float,
                 state_file="active_positions.json", excel_path="trading_log.xlsx"):
        self.per_trade_cash = per_trade_cash
        self.max_open = max_open
        self.budget = budget
        self.state_file = state_file
        self.excel_path = excel_path
        self._load()
        # preparar excel (incluye fuente precio y hoja skips)
        ensure_excel_with_headers(
            self.excel_path,
            "opens",
            [
                "opened_at","opened_date","opened_time","opened_minute",
                "id","ticker","ticker_yf","side",
                "entry_price","price_source_open","qty","investment",
                "sl","sl_pct","tp","tp_pct","rr",
                "prob","atr_used","cash_available_post_open","month"
            ],
        )
        ensure_excel_with_headers(
            self.excel_path,
            "closes",
            [
                "closed_at","closed_date","closed_time","closed_minute",
                "id","ticker","ticker_yf","side",
                "entry_price","exit_price","price_source_close","qty","investment",
                "pnl","pnl_pct","result","held_days",
                "sl","tp","rr","prob","atr_used","month"
            ],
        )
        ensure_skips_sheet(self.excel_path)

    def _load(self):
        if os.path.exists(self.state_file):
            with open(self.state_file, "r", encoding="utf-8") as f:
                self.positions = json.load(f)
            # ---- MIGRACIÓN DE CAMPOS FALTANTES ----
            changed = False
            for pid, p in list(self.positions.items()):
                if "ticker_yf" not in p and "ticker" in p:
                    p["ticker_yf"] = ensure_mx(p["ticker"]); changed = True
                if "rr" not in p: p["rr"] = None; changed = True
                if "prob" not in p: p["prob"] = None; changed = True
                if "sl_pct" not in p and "entry_price" in p and "sl" in p:
                    try:
                        p["sl_pct"] = max(0.0, (p["entry_price"] - p["sl"]) / p["entry_price"]); changed = True
                    except Exception:
                        p["sl_pct"] = None; changed = True
                if "tp_pct" not in p and "entry_price" in p and "tp" in p:
                    try:
                        p["tp_pct"] = max(0.0, (p["tp"] - p["entry_price"]) / p["entry_price"]); changed = True
                    except Exception:
                        p["tp_pct"] = None; changed = True
                if "atr_used" not in p: p["atr_used"] = None; changed = True
                if "month" not in p: p["month"] = ""; changed = True
            if changed:
                with open(self.state_file, "w", encoding="utf-8") as f:
                    json.dump(self.positions, f, indent=2, ensure_ascii=False)
        else:
            self.positions = {}

    def _save(self):
        with open(self.state_file, "w", encoding="utf-8") as f:
            json.dump(self.positions, f, indent=2, ensure_ascii=False)

    def active_count(self):
        return sum(1 for p in self.positions.values() if p["status"] == "ACTIVE")

    def used_budget(self):
        return sum(p["investment"] for p in self.positions.values() if p["status"] == "ACTIVE")

    # ==== Excel logging helpers ====
    @staticmethod
    def _split_dt_fields(dt_iso: str):
        dt = datetime.fromisoformat(dt_iso)
        return dt.date().isoformat(), dt.strftime("%H:%M:%S"), f"{dt.minute:02d}"

    def log_open_excel(self, p: dict):
        opened_date, opened_time, opened_minute = self._split_dt_fields(p["opened_at"])
        append_excel_row(
            self.excel_path,
            "opens",
            [
                p["opened_at"], opened_date, opened_time, opened_minute,
                p["id"], p["ticker"], get_ticker_yf_from_pos(p), p["side"],
                p["entry_price"], p.get("price_source_open",""), p["qty"], p["investment"],
                p["sl"], p["sl_pct"], p["tp"], p["tp_pct"], p["rr"],
                p.get("prob", None), p.get("atr_used", None), p.get("cash_available_post_open", None), p.get("month","")
            ]
        )

    def log_close_excel(self, p: dict, pnl: float, pnl_pct: float, held_days: int, price_source_close: str = ""):
        closed_date, closed_time, closed_minute = self._split_dt_fields(p["exit_at"])
        append_excel_row(
            self.excel_path,
            "closes",
            [
                p["exit_at"], closed_date, closed_time, closed_minute,
                p["id"], p["ticker"], get_ticker_yf_from_pos(p), p["side"],
                p["entry_price"], p["exit_price"], price_source_close, p["qty"], p["investment"],
                pnl, pnl_pct, p.get("result",""), held_days,
                p["sl"], p["tp"], p.get("rr", None), p.get("prob", None), p.get("atr_used", None), p.get("month","")
            ]
        )

    # ==== Apertura/Cierre ====
    def try_open_buy(self, ticker: str, price: float, price_source_open: str,
                     sl: float, tp: float, horizon_days: int,
                     per_trade_cash: float, prob, rr: float, month: str,
                     sl_pct: float, tp_pct: float, atr_used: float, ticker_yf: str):
        if self.active_count() >= self.max_open:
            return None, "MAX_OPEN_REACHED"
        if self.used_budget() + per_trade_cash > self.budget:
            return None, "BUDGET_EXCEEDED"

        qty = max(1, int(per_trade_cash // price))
        invest = qty * price

        pid = f"{ticker}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.positions[pid] = {
            "id": pid,
            "ticker": ticker,
            "ticker_yf": ticker_yf,
            "side": "BUY",
            "entry_price": price,
            "price_source_open": price_source_open,
            "tp": tp,
            "sl": sl,
            "sl_pct": sl_pct,
            "tp_pct": tp_pct,
            "qty": qty,
            "investment": invest,
            "opened_at": datetime.now().isoformat(),
            "max_holding_days": int(horizon_days),
            "status": "ACTIVE",
            "prob": prob,
            "rr": rr,
            "atr_used": atr_used,
            "month": month
        }
        self._save()
        self.positions[pid]["cash_available_post_open"] = self.budget - self.used_budget()
        self._save()
        self.log_open_excel(self.positions[pid])
        return pid, "OPENED"

    def _force_close(self, pid: str, last_price: float, reason: str, price_source_close: str = ""):
        p = self.positions[pid]
        qty = p["qty"]; entry = p["entry_price"]
        pnl = (last_price - entry) * qty
        p["status"] = "CLOSED"
        p["exit_price"] = last_price
        p["exit_at"] = datetime.now().isoformat()
        p["pnl"] = pnl
        p["result"] = reason
        self._save()
        held_days = (datetime.fromisoformat(p["exit_at"]) - datetime.fromisoformat(p["opened_at"])).days
        pnl_pct = pnl / max(1e-9, p["investment"])
        self.log_close_excel(p, pnl, pnl_pct, held_days, price_source_close=price_source_close)
        return pnl, pnl_pct, held_days

    def try_close_hits(self, pid: str, last_price: float, price_source_close: str = ""):
        p = self.positions[pid]
        if p["status"] != "ACTIVE": return False, None, 0.0, 0.0, 0
        tp = p["tp"]; sl = p["sl"]; entry = p["entry_price"]
        hit = None
        if last_price >= tp: hit = "TP_HIT"
        elif last_price <= sl: hit = "SL_HIT"

        if hit:
            pnl, pnl_pct, held_days = self._force_close(pid, last_price, hit, price_source_close=price_source_close)
            return True, hit, pnl, pnl_pct, held_days

        opened = datetime.fromisoformat(p["opened_at"])
        if datetime.now() >= opened + timedelta(days=int(p["max_holding_days"])):
            pnl, pnl_pct, held_days = self._force_close(pid, last_price, "TIME_EXIT", price_source_close=price_source_close)
            return True, "TIME_EXIT", pnl, pnl_pct, held_days

        return False, None, 0.0, 0.0, 0

# ========= util cierre masivo =========
def close_all_active_now(broker: PaperBroker):
    active = [(pid, p) for pid, p in broker.positions.items() if p["status"] == "ACTIVE"]
    if not active:
        print("No hay posiciones activas para cerrar.")
        return 0, 0.0

    tickers = list({ensure_mx(p["ticker"]) for _, p in active})
    prices_dict = fetch_prices_multi(tickers, prefer="yf")

    closed_n = 0
    total_pnl = 0.0
    for pid, p in active:
        tmx = ensure_mx(p["ticker"])
        if tmx not in prices_dict:
            print(f"[RESET] Sin precio para {tmx}, salto cierre forzado.")
            continue
        last_price, source = prices_dict[tmx]
        pnl, pnl_pct, held_days = broker._force_close(pid, last_price, "MANUAL_RESET", price_source_close=source)
        closed_n += 1
        total_pnl += pnl
        rr = p.get("rr", None); rr_txt = f"{rr:.2f}x" if rr is not None else "-"
        prob_txt = prob_to_pct(p.get("prob", None))
        send_tg(
            f"⚠️ <b>CIERRE MASIVO</b>\n"
            f"📈 <b>{p['ticker']}</b> ({get_ticker_yf_from_pos(p)}) BUY  |  Qty: <b>{p['qty']}</b>\n"
            f"📍 MANUAL_RESET  |  ⏳ {held_days}d  |  R≈{rr_txt}\n"
            f"💰 Entrada: {pretty_cash(p['entry_price'])}  →  Salida: {pretty_cash(last_price)}\n"
            f"🛰️ Fuente precio: {source}\n"
            f"📊 P&L: <b>{pretty_cash(pnl)}</b> ({pnl_pct*100:+.1f}%)\n"
            f"🔮 Prob. éxito (señal): {prob_txt}"
        )
    print(f"[RESET] Cerradas {closed_n} posiciones. PnL total ~ {pretty_cash(total_pnl)}")
    return closed_n, total_pnl

# ========= Monitor =========
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--month", required=True, help="Mes objetivo YYYY-MM (ej. 2025-10)")
    parser.add_argument("--base-dir", default=".", help="Raíz del repo (default: .)")
    parser.add_argument("--interval-min", type=int, default=15, help="Minutos entre chequeos (default 15)")
    parser.add_argument("--max-new", type=int, default=3, help="Máximo de nuevas aperturas por ciclo")

    # Control dinámico de SL y R:R solo BUY
    parser.add_argument("--only-buy", action="store_true", default=True, help="Forzar solo BUY")
    parser.add_argument("--atr-mult", type=float, default=4.0, help="Multiplicador de ATR para SL dinámico (default 4.0)")
    parser.add_argument("--min-rr", type=float, default=4.0, help="Relación mínima R:R para entrar (default 4.0)")
    parser.add_argument("--fallback-sl-pct", type=float, default=0.001, help="SL porcentual de respaldo si no hay ATR (default 0.1%)")
    parser.add_argument("--tp-ceiling-atr-mult", type=float, default=5.0,
                        help="Techo opcional: máx TP = entry + (este_mult * ATR). Ej: 5.0")
    parser.add_argument("--excel-path", type=str, default="trading_log.xlsx", help="Ruta del Excel de bitácora")

    # Reset al inicio
    parser.add_argument("--ask-reset", action="store_true", default=True,
                        help="Preguntar al iniciar si deseas cerrar todas las posiciones activas.")
    parser.add_argument("--force-reset", action="store_true", default=False,
                        help="Cerrar todas las posiciones activas al precio actual sin preguntar.")
    parser.add_argument("--no-ask-reset", action="store_true", default=False,
                        help="No preguntar por reset al inicio (ignora --ask-reset).")

    # Novedades: logging de descartes
    parser.add_argument("--log-all-candidates", action="store_true", default=False,
                        help="Registrar en hoja 'skips' todos los candidatos que no se abren (ranking/capacidad/etc.).")
    parser.add_argument("--skips-to-telegram", action="store_true", default=False,
                        help="Enviar un resumen de descartes por motivo al final de cada ciclo.")

    # Preferencia de fuente
    parser.add_argument("--prefer", choices=["yf","investing"], default="yf",
                        help="Fuente preferida para precios (default yf)")

    args = parser.parse_args()
    month = args.month
    base_dir = os.path.abspath(args.base_dir)

    # Política base
    policy = load_policy(month, base_dir)
    tp_pct_base = float(policy["tp_pct"] or 0.0)
    sl_pct_fallback = float(policy["sl_pct"] or args.fallback_sl_pct)
    horizon_days = int(policy["horizon_days"])
    per_trade_cash = float(policy["per_trade_cash"])
    max_open = int(policy["max_open"])
    budget = float(policy["budget"])

    print("=== Paper monitor BMV (solo BUY, SL ATR, TP híbrido, Excel log v4 + multi-fuente + skips) ===")
    print(f"Mes: {month}")
    print(f"H={horizon_days}d  cash/trade={pretty_cash(per_trade_cash)}  max_open={max_open}  budget={pretty_cash(budget)}")
    print(f"SL dinámico: ATR*{args.atr_mult} (respaldo {sl_pct_fallback*100:.2f}%), min R:R {args.min_rr:.1f}x, tp_pct_política={tp_pct_base*100:.1f}%")
    if args.tp_ceiling_atr_mult is not None:
        print(f"Techo TP: entry + {args.tp_ceiling_atr_mult}*ATR")
    print(f"Excel de bitácora: {args.excel_path}")

    send_tg(
        f"🚀 <b>MONITOR BMV</b> (solo BUY, SL=ATR*{args.atr_mult}, TP híbrido, RR≥{args.min_rr}x)\n"
        f"🗓 Mes {month}\n"
        f"⏳ H {horizon_days}d | 💵 {pretty_cash(per_trade_cash)} por trade | 📊 max_open={max_open} | 💼 budget={pretty_cash(budget)}"
    )

    # Broker
    broker = PaperBroker(
        per_trade_cash=per_trade_cash,
        max_open=max_open,
        budget=budget,
        state_file="active_positions.json",
        excel_path=args.excel_path
    )

    # === Reset al iniciar (opcional) ===
    if args.force_reset:
        print("[RESET] Forzado por flag --force-reset.")
        close_all_active_now(broker)
    elif not args.no_ask_reset and args.ask_reset:
        if broker.active_count() > 0:
            try:
                resp = input("¿Cerrar TODAS las posiciones activas ahora para reiniciar? [s/N]: ").strip().lower()
            except EOFError:
                resp = "n"
            if resp in ("s","si","sí","y","yes"):
                close_all_active_now(broker)
            else:
                print("[RESET] No se cerraron posiciones al inicio.")

    # Forecast → solo BUY + top score
    df_fc = load_forecast(month, base_dir)
    today = datetime.now().date()
    cand = pick_today_signals(df_fc, today, max_candidates=50, only_buy=args.only_buy)

    if cand.empty:
        print("No hay señales BUY para hoy (o no hay fecha <= hoy en el forecast).")
        send_tg("ℹ️ No hay señales BUY para hoy en el forecast.")
        return

    print(f"Señales BUY candidatas hoy: {len(cand)} (top por score)")
    print(cand.head(10).to_string(index=False))

    while True:
        try:
            # 1) Cerrar por TP/SL/horizonte
            active = [(pid, p) for pid, p in broker.positions.items() if p["status"] == "ACTIVE"]
            if active:
                print(f"[{datetime.now():%H:%M:%S}] Activas: {len(active)}")
                tickers = list({ensure_mx(p["ticker"]) for _, p in active})
                if tickers:
                    price_dict = fetch_prices_multi(tickers, prefer=args.prefer)
                    for pid, p in active:
                        tmx = ensure_mx(p["ticker"])
                        if tmx in price_dict:
                            last_price, source = price_dict[tmx]
                            closed, reason, pnl, pnl_pct, held_days = broker.try_close_hits(pid, last_price, price_source_close=source)
                            if closed:
                                emoji = "🎉" if reason == "TP_HIT" else ("🛑" if reason == "SL_HIT" else "⏰")
                                entry = p["entry_price"]; exit_px = broker.positions[pid]["exit_price"]
                                qty = p["qty"]; inv = max(1e-9, p["investment"])
                                rr = p.get("rr", None); rr_txt = f"{rr:.2f}x" if rr is not None else "-"
                                prob_txt = prob_to_pct(p.get("prob", None))
                                send_tg(
                                    f"{emoji} <b>POSICIÓN CERRADA</b>\n"
                                    f"📈 <b>{p['ticker']}</b> ({get_ticker_yf_from_pos(p)}) BUY  |  Qty: <b>{qty}</b>\n"
                                    f"📍 {reason}  |  ⏳ {held_days}d  |  R≈{rr_txt}\n"
                                    f"💰 Entrada: {pretty_cash(entry)}  →  Salida: {pretty_cash(exit_px)}\n"
                                    f"🛰️ Fuente precio: {source}\n"
                                    f"🧮 Inversión: {pretty_cash(inv)}\n"
                                    f"📊 P&L: <b>{pretty_cash(pnl)}</b> ({pnl_pct*100:+.1f}%)\n"
                                    f"🔮 Prob. éxito (señal): {prob_txt}"
                                )
                        else:
                            print(f"Sin precio para {tmx} en este ciclo (ni YF ni Investing).")
                            # Log opcional de skip al cerrar
                            log_skip_excel(
                                broker.excel_path, month, p["ticker"], tmx, p.get("side","BUY"),
                                reason_code="no_price_close",
                                reason_msg="No hay precio para evaluar cierre",
                                rr=p.get("rr", None), min_rr=args.min_rr, atr_used=p.get("atr_used", None),
                                sl_pct=p.get("sl_pct", None), tp_pct=p.get("tp_pct", None),
                                entry_price=p.get("entry_price", None), price_source=None,
                                per_trade_cash=broker.per_trade_cash,
                                budget_used=broker.used_budget(), budget_total=broker.budget,
                                max_open=max_open, active_count=broker.active_count()
                            )

            # 2) Abrir nuevas si hay capacidad (solo BUY)
            capacity = max(0, max_open - broker.active_count())
            to_open = min(capacity, args.max_new)

            # Si no hay capacidad, loggea candidatos pendientes y duerme
            if to_open <= 0 and args.log_all_candidates:
                active_tickers = {p["ticker"] for p in broker.positions.values() if p["status"] == "ACTIVE"}
                pending = cand[~cand["ticker"].isin(active_tickers)]
                for _, row in pending.iterrows():
                    t = row["ticker"]; tmx = ensure_mx(t)
                    log_skip_excel(
                        broker.excel_path, month, t, tmx, "BUY",
                        reason_code="capacity_full",
                        reason_msg=f"Sin capacidad: active={broker.active_count()} / max_open={max_open}",
                        rr=None, min_rr=args.min_rr, atr_used=None, sl_pct=None, tp_pct=None,
                        entry_price=None, price_source=None,
                        per_trade_cash=broker.per_trade_cash,
                        budget_used=broker.used_budget(), budget_total=broker.budget,
                        max_open=max_open, active_count=broker.active_count()
                    )

            if to_open > 0:
                active_tickers = {p["ticker"] for p in broker.positions.values() if p["status"] == "ACTIVE"}
                open_candidates = cand[~cand["ticker"].isin(active_tickers)].head(to_open)

                # Log los descartados por ya estar activos (si pides log-all)
                if args.log_all_candidates:
                    skipped_active = cand[cand["ticker"].isin(active_tickers)]
                    for _, row in skipped_active.iterrows():
                        t = row["ticker"]; tmx = ensure_mx(t)
                        log_skip_excel(
                            broker.excel_path, month, t, tmx, "BUY",
                            reason_code="already_active",
                            reason_msg="Ticker ya tiene posición activa",
                            rr=None, min_rr=args.min_rr, atr_used=None, sl_pct=None, tp_pct=None,
                            entry_price=None, price_source=None,
                            per_trade_cash=broker.per_trade_cash,
                            budget_used=broker.used_budget(), budget_total=broker.budget,
                            max_open=max_open, active_count=broker.active_count()
                        )

                if not open_candidates.empty:
                    tickers_new_mx = [ensure_mx(t) for t in open_candidates["ticker"].tolist()]
                    price_dict_open = fetch_prices_multi(tickers_new_mx, prefer=args.prefer)

                    opened_tickers: List[str] = []
                    skipped_this_cycle: List[Tuple[str, str]] = []  # (ticker_mx, reason_code)

                    for _, row in open_candidates.iterrows():
                        t = row["ticker"]
                        tmx = ensure_mx(t)
                        if tmx not in price_dict_open:
                            print(f"[OPEN] {tmx} descartado: sin precio (ni YF ni Investing).")
                            if args.log_all_candidates:
                                log_skip_excel(
                                    broker.excel_path, month, t, tmx, "BUY",
                                    reason_code="no_price",
                                    reason_msg="No hay precio en YF ni Investing (ni ffill)",
                                    rr=None, min_rr=args.min_rr, atr_used=None, sl_pct=None, tp_pct=None,
                                    entry_price=None, price_source=None,
                                    per_trade_cash=broker.per_trade_cash,
                                    budget_used=broker.used_budget(), budget_total=broker.budget,
                                    max_open=max_open, active_count=broker.active_count()
                                )
                            skipped_this_cycle.append((tmx, "no_price"))
                            continue

                        price, price_source = price_dict_open[tmx]

                        atr_val = compute_atr_14(tmx, lookback_days=60)
                        sl, tp, risk_ps, rr, tp_pct_vs_entry, sl_pct_vs_entry = compute_tp_sl_buy_hybrid(
                            entry=price,
                            atr=atr_val,
                            atr_mult=args.atr_mult,
                            fallback_sl_pct=sl_pct_fallback,
                            min_rr=args.min_rr,
                            tp_pct_policy=tp_pct_base,
                            tp_ceiling_atr_mult=args.tp_ceiling_atr_mult
                        )

                        if rr < args.min_rr:
                            print(f"[OPEN] {tmx} descartado: R {rr:.2f} < min_rr {args.min_rr:.2f} | "
                                  f"SL%={sl_pct_vs_entry*100:.2f} TP%={tp_pct_vs_entry*100:.2f} ATR={atr_val}")
                            if args.log_all_candidates:
                                log_skip_excel(
                                    broker.excel_path, month, t, tmx, "BUY",
                                    reason_code="rr_too_low",
                                    reason_msg=f"R {rr:.2f} < min_rr {args.min_rr:.2f}",
                                    rr=rr, min_rr=args.min_rr, atr_used=atr_val,
                                    sl_pct=sl_pct_vs_entry, tp_pct=tp_pct_vs_entry,
                                    entry_price=price, price_source=price_source,
                                    per_trade_cash=broker.per_trade_cash,
                                    budget_used=broker.used_budget(), budget_total=broker.budget,
                                    max_open=max_open, active_count=broker.active_count()
                                )
                            skipped_this_cycle.append((tmx, "rr_too_low"))
                            continue

                        pid, status = broker.try_open_buy(
                            ticker=t, price=price, price_source_open=price_source,
                            sl=sl, tp=tp,
                            horizon_days=horizon_days, per_trade_cash=per_trade_cash,
                            prob=float(row.get("score", 0.0)),
                            rr=rr, month=month,
                            sl_pct=sl_pct_vs_entry, tp_pct=tp_pct_vs_entry,
                            atr_used=float(atr_val) if atr_val is not None else None,
                            ticker_yf=tmx
                        )
                        if status == "OPENED":
                            opened_tickers.append(tmx)
                            p = broker.positions[pid]
                            qty = p["qty"]
                            invest = p["investment"]
                            pnl_target = (tp - price) * qty
                            prob_txt = prob_to_pct(p.get("prob", None))
                            send_tg(
                                f"🚀 <b>NUEVA POSICIÓN</b>\n"
                                f"📈 <b>{t}</b> ({tmx}) BUY\n"
                                f"💰 Entrada: {pretty_cash(price)}  |  Qty: <b>{qty}</b>\n"
                                f"🧮 Inversión: {pretty_cash(invest)}  |  💼 Cash post-apertura: {pretty_cash(p.get('cash_available_post_open', 0.0))}\n"
                                f"🛡️ SL: {pretty_cash(sl)}  ({sl_pct_vs_entry*100:.2f}%)\n"
                                f"🎯 TP: {pretty_cash(tp)}  ({tp_pct_vs_entry*100:.2f}%)\n"
                                f"📏 ATR usado: {pretty_cash(p.get('atr_used', 0.0) or 0.0)}\n"
                                f"🛰️ Fuente precio: {price_source}\n"
                                f"🔢 R≈{rr:.2f}x  |  🔮 Prob. éxito: {prob_txt}\n"
                                f"📊 Ganancia estimada a TP: <b>{pretty_cash(pnl_target)}</b>\n"
                                f"⏳ Horizonte: {horizon_days}d"
                            )
                        else:
                            reason = "max_open_reached" if status == "MAX_OPEN_REACHED" else (
                                     "budget_exceeded" if status == "BUDGET_EXCEEDED" else "other")
                            if args.log_all_candidates:
                                log_skip_excel(
                                    broker.excel_path, month, t, tmx, "BUY",
                                    reason_code=reason,
                                    reason_msg=f"status={status}",
                                    rr=rr, min_rr=args.min_rr, atr_used=atr_val,
                                    sl_pct=sl_pct_vs_entry, tp_pct=tp_pct_vs_entry,
                                    entry_price=price, price_source=price_source,
                                    per_trade_cash=broker.per_trade_cash,
                                    budget_used=broker.used_budget(), budget_total=broker.budget,
                                    max_open=max_open, active_count=broker.active_count()
                                )
                            skipped_this_cycle.append((tmx, reason))
                            print(f"No abrí {t}: {status}")

                    # Si pediste resumen a Telegram de los skips del ciclo
                    if args.skips_to_telegram and skipped_this_cycle:
                        cnt = Counter([r for _, r in skipped_this_cycle])
                        detalle = "\n".join([f"• {k}: {v}" for k, v in cnt.items()])
                        send_tg(
                            f"ℹ️ <b>Descartes en este ciclo</b>\n"
                            f"{detalle}"
                        )

                # Si pediste log-all y sobraron candidatos por ranking/capacidad (no llegaron a evaluar)
                if args.log_all_candidates:
                    evaluated = set([*open_candidates["ticker"].tolist(), *opened_tickers]) if not open_candidates.empty else set()
                    # todo el set de candidatos menos activos menos evaluados
                    active_tickers = {p["ticker"] for p in broker.positions.values() if p["status"] == "ACTIVE"}
                    leftover = cand[~cand["ticker"].isin(active_tickers | evaluated)]
                    for _, row in leftover.iterrows():
                        t = row["ticker"]; tmx = ensure_mx(t)
                        log_skip_excel(
                            broker.excel_path, month, t, tmx, "BUY",
                            reason_code="not_evaluated_capacity",
                            reason_msg="No se evaluó por límite de capacidad/orden de ranking",
                            rr=None, min_rr=args.min_rr, atr_used=None, sl_pct=None, tp_pct=None,
                            entry_price=None, price_source=None,
                            per_trade_cash=broker.per_trade_cash,
                            budget_used=broker.used_budget(), budget_total=broker.budget,
                            max_open=max_open, active_count=broker.active_count()
                        )

            # 3) Espera
            wake = (datetime.now()+timedelta(minutes=args.interval_min)).strftime("%H:%M")
            print(f"Duerme {args.interval_min} min - {wake}")
            time.sleep(args.interval_min * 60)

        except KeyboardInterrupt:
            print("Detenido por usuario.")
            send_tg("🛑 Monitor BMV detenido por el usuario.")
            break
        except Exception as e:
            print(f"Error ciclo: {e}")
            time.sleep(10)

if __name__ == "__main__":
    main()
