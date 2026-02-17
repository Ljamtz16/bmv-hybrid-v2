# monitor_bmv_buy_atr_rr_hybrid_excel_v7.py
import os, json, time, argparse
from datetime import datetime, timedelta, date
from typing import Dict, Tuple, List, Optional
import pandas as pd
import numpy as np
import yfinance as yf
import requests
from dotenv import load_dotenv, find_dotenv
from openpyxl import Workbook, load_workbook

# ================== Config / Telegram ==================
load_dotenv()
env_path = find_dotenv(usecwd=True)
loaded = load_dotenv(env_path, override=True)
print("ENV loaded:", loaded, "| path:", env_path or "<none>")

TOKEN = (os.getenv("TELEGRAM_BOT_TOKEN") or "").strip()
CHAT_ID = (os.getenv("TELEGRAM_CHAT_ID") or "").strip()

def send_tg(msg: str) -> bool:
    if not TOKEN or not CHAT_ID:
        print(f"[TG] Faltan credenciales: TOKEN={'OK' if TOKEN else 'MISSING'} CHAT_ID={CHAT_ID}")
        return False
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{TOKEN}/sendMessage",
            data={"chat_id": CHAT_ID, "text": msg, "parse_mode": "HTML"},
            timeout=15,
        )
        ok = r.status_code == 200 and r.json().get("ok", False)
        if not ok:
            print(f"[TG] Error: HTTP {r.status_code} {r.text}")
        return ok
    except Exception as e:
        print(f"[TG] Excepción: {e}")
        return False

# ================== Utils ==================
def read_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def safe_float(x, default=None):
    try:
        if x is None: return default
        if isinstance(x, (int, float)): return float(x)
        return float(str(x).replace(",", ""))
    except:
        return default

def ensure_mx(ticker: str) -> str:
    t = (ticker or "").strip()
    return t if t.endswith(".MX") else f"{t}.MX"

def latest_business_day_in(df_dates: pd.Series, until: date) -> Optional[pd.Timestamp]:
    d = pd.to_datetime(df_dates).dt.date
    d = d[d <= until]
    if d.empty: return None
    return pd.Timestamp(max(d))

def pretty_cash(x: float) -> str:
    try:
        return f"${x:,.2f}"
    except:
        return str(x)

def prob_to_pct(prob):
    try:
        p = float(prob)
        if 0.0 <= p <= 1.0:
            return f"{p*100:.1f}%"
        if 1.0 < p <= 100.0:
            return f"{p:.1f}%"
        return f"{p:.3f}"
    except:
        return "-"

def get_ticker_yf_from_pos(p: dict) -> str:
    return p.get("ticker_yf") or ensure_mx(p.get("ticker",""))

# ================== Excel helpers ==================
def ensure_excel_with_headers(path: str, sheet_name: str, headers: List[str]):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        wb.save(path)
        return
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        wb.save(path)
        return
    ws = wb[sheet_name]
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(headers)
        wb.save(path)

def append_excel_row(path: str, sheet_name: str, row_values: List):
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    ws = wb[sheet_name]
    ws.append(row_values)
    wb.save(path)

# ================== Política ==================
def load_policy(month: str, base_dir: str):
    wfjson = os.path.join(base_dir, "reports", "forecast", month, "policy_wfbox.json")
    if os.path.exists(wfjson):
        p = read_json(wfjson)
        return {
            "tp_pct": safe_float(p.get("tp_pct")),
            "sl_pct": safe_float(p.get("sl_pct")),
            "horizon_days": int(p.get("horizon_days", 5)),
            "per_trade_cash": safe_float(p.get("per_trade_cash", 2000)),
            "max_open": int(p.get("max_open", 5)),
            "budget": safe_float(p.get("budget", 10000)),
        }
    wf_csv = os.path.join(base_dir, "wf_box", "reports", "forecast", "policy_selected_walkforward.csv")
    if not os.path.exists(wf_csv):
        raise FileNotFoundError("No hay política: faltan policy_wfbox.json y wf_box/.../policy_selected_walkforward.csv")
    df = pd.read_csv(wf_csv)
    if "month" not in df.columns:
        raise ValueError("El CSV de walk-forward no tiene columna 'month'")
    row = df.loc[df["month"] == month]
    if row.empty:
        raise ValueError(f"No encontré política para el mes {month} en {wf_csv}")
    r = row.iloc[0].to_dict()
    return {
        "tp_pct": safe_float(r.get("tp_pct")),
        "sl_pct": safe_float(r.get("sl_pct")),
        "horizon_days": int(safe_float(r.get("horizon_days"), 5)),
        "per_trade_cash": safe_float(r.get("per_trade_cash", 2000)),
        "max_open": int(safe_float(r.get("max_open"), 5)),
        "budget": safe_float(r.get("budget", 10000)),
    }

# ================== Forecast ==================
def load_forecast(month: str, base_dir: str) -> pd.DataFrame:
    fc_path = os.path.join(base_dir, "reports", "forecast", month, f"forecast_{month}_with_gate.csv")
    if not os.path.exists(fc_path):
        raise FileNotFoundError(f"No existe {fc_path}. Genera el forecast primero.")
    df = pd.read_csv(fc_path)

    if "ticker" not in df.columns:
        if "symbol" in df.columns: df = df.rename(columns={"symbol":"ticker"})
        else: raise ValueError("El forecast no tiene columna 'ticker' (ni 'symbol').")

    date_col = None
    for c in ["date", "dt", "signal_date", "session_date"]:
        if c in df.columns:
            date_col = c; break
    if date_col is None: raise ValueError("No encontré columna de fecha (date/dt/signal_date).")
    df["date"] = pd.to_datetime(df[date_col], errors="coerce")

    if "side" not in df.columns:
        for c in ["signal", "action"]:
            if c in df.columns:
                df["side"] = df[c].str.upper()
                break
        if "side" not in df.columns:
            raise ValueError("No encontré columna 'side'/'signal'/'action' (BUY/SELL).")

    score_col = None
    for c in ["prob", "abs_y", "score", "yhat_prob"]:
        if c in df.columns:
            score_col = c; break
    if score_col is None:
        df["__score__"] = 1.0
        score_col = "__score__"

    df["score"] = pd.to_numeric(df[score_col], errors="coerce").fillna(0.0)
    return df[["ticker","date","side","score"]].copy()

def pick_recent_signals(df: pd.DataFrame, today: date, n_recent_days: int = 3, only_buy: bool = True):
    df = df.copy()
    df = df[df["date"].dt.date <= today]
    if df.empty:
        return df.iloc[0:0].copy(), []
    if only_buy:
        df = df[df["side"].str.upper() == "BUY"].copy()
    if df.empty:
        return df.iloc[0:0].copy(), []

    unique_dates = sorted(df["date"].dt.date.unique())
    pick_dates = unique_dates[-n_recent_days:] if len(unique_dates) >= n_recent_days else unique_dates
    out = df[df["date"].dt.date.isin(pick_dates)].copy()

    out = out.sort_values(["ticker","score"], ascending=[True, False])
    out = out.groupby("ticker", as_index=False).first()
    out = out.sort_values("score", ascending=False).reset_index(drop=True)
    return out, pick_dates

# ================== Investing (stub) ==================
INVESTING_MAP: Dict[str, Tuple[str, str]] = {}

def investing_last_close_dict(tickers_mx: List[str]) -> Tuple[Dict[str, Tuple[float, str]], Dict[str, str]]:
    # Conecta tu scraper/API aquí; por ahora deja stub.
    res: Dict[str, Tuple[float, str]] = {}
    diag: Dict[str, str] = {t:"inv_stub" for t in tickers_mx}
    return res, diag

# ================== YF price helpers (daily + intraday) ==================
def yf_last_close_dict(tickers_mx: List[str]) -> Tuple[Dict[str, Tuple[float, str]], Dict[str, str]]:
    res: Dict[str, Tuple[float, str]] = {}
    diag: Dict[str, str] = {}
    if not tickers_mx: return res, diag
    try:
        data = yf.download(tickers=tickers_mx, period="2d", interval="1d",
                           progress=False, group_by="ticker", threads=True, auto_adjust=False)
        for t in tickers_mx:
            try:
                s = data["Close"].dropna() if len(tickers_mx) == 1 else data[t]["Close"].dropna()
                if not s.empty:
                    res[t] = (float(s.iloc[-1]), "yf")
                    diag[t] = (diag.get(t,"") + "|yf_ok").strip("|")
                else:
                    diag[t] = (diag.get(t,"") + "|yf_empty").strip("|")
            except Exception as e:
                diag[t] = (diag.get(t,"") + f"|yf_error:{e}").strip("|")
    except Exception as e:
        for t in tickers_mx:
            diag[t] = (diag.get(t,"") + f"|yf_batch_error:{e}").strip("|")
    return res, diag

def yf_intraday_dict(tickers_mx: List[str], interval: str = "1m", window_min: int = 15) -> Tuple[Dict[str, Tuple[float, str]], Dict[str, str]]:
    """
    Toma el último precio válido dentro de los últimos `window_min` minutos en intradía.
    """
    res: Dict[str, Tuple[float, str]] = {}
    diag: Dict[str, str] = {}
    if not tickers_mx: return res, diag
    try:
        # Para múltiples tickers, yfinance devuelve multiindex
        data = yf.download(tickers=tickers_mx, period="1d", interval=interval,
                           progress=False, group_by="ticker", threads=True, auto_adjust=False)
        now = pd.Timestamp.now(tz=None)
        cutoff = now - pd.Timedelta(minutes=window_min)

        for t in tickers_mx:
            try:
                df = data if len(tickers_mx) == 1 else data[t]
                s = df["Close"].dropna()
                if s.empty:
                    diag[t] = (diag.get(t,"") + "|intraday_empty").strip("|")
                    continue
                # último dentro de la ventana
                s_win = s[s.index >= cutoff]
                if s_win.empty:
                    # tomar el último del día (puede estar justo fuera del cutoff)
                    price = float(s.iloc[-1])
                    res[t] = (price, f"yf-intra-{interval}")
                    diag[t] = (diag.get(t,"") + "|intraday_old_but_used").strip("|")
                else:
                    price = float(s_win.iloc[-1])
                    res[t] = (price, f"yf-intra-{interval}")
                    diag[t] = (diag.get(t,"") + "|intraday_ok").strip("|")
            except Exception as e:
                diag[t] = (diag.get(t,"") + f"|intraday_error:{e}").strip("|")
    except Exception as e:
        for t in tickers_mx:
            diag[t] = (diag.get(t,"") + f"|intraday_batch_error:{e}").strip("|")
    return res, diag

def fetch_prices_multi(tickers_mx: List[str], prefer: str = "yf") -> Tuple[Dict[str, Tuple[float, str]], Dict[str, str]]:
    # Diario + fallbacks
    res: Dict[str, Tuple[float, str]] = {}
    diag: Dict[str, str] = {}
    if not tickers_mx: return res, diag

    if prefer == "yf":
        a, da = yf_last_close_dict(tickers_mx); res.update(a); diag.update(da)
        need = [t for t in tickers_mx if t not in res]
        b, db = investing_last_close_dict(need); res.update(b)
        for k,v in db.items(): diag[k] = (diag.get(k,"") + "|" + v).strip("|")
    else:
        a, da = investing_last_close_dict(tickers_mx); res.update(a); diag.update(da)
        need = [t for t in tickers_mx if t not in res]
        b, db = yf_last_close_dict(need); res.update(b)
        for k,v in db.items(): diag[k] = (diag.get(k,"") + "|" + v).strip("|")

    # ffill 30d
    missing = [t for t in tickers_mx if t not in res]
    if missing:
        try:
            data = yf.download(tickers=missing, period="30d", interval="1d",
                               progress=False, group_by="ticker", threads=True, auto_adjust=False)
            for t in missing:
                try:
                    s = data["Close"].ffill() if len(missing) == 1 else data[t]["Close"].ffill()
                    if not s.dropna().empty:
                        res[t] = (float(s.dropna().iloc[-1]), "yf-ffill")
                        diag[t] = (diag.get(t,"") + "|ffill_ok").strip("|")
                    else:
                        diag[t] = (diag.get(t,"") + "|ffill_empty").strip("|")
                except Exception as e:
                    diag[t] = (diag.get(t,"") + f"|ffill_error:{e}").strip("|")
        except Exception as e:
            for t in missing:
                diag[t] = (diag.get(t,"") + f"|ffill_batch_error:{e}").strip("|")

    # Ultra fallback
    missing = [t for t in tickers_mx if t not in res]
    for t in missing:
        try:
            tk = yf.Ticker(t)
            h = tk.history(period="5d", interval="1d", auto_adjust=False)
            if h is not None and not h.empty and "Close" in h.columns:
                val = float(h["Close"].dropna().iloc[-1])
                res[t] = (val, "yf-ticker-history")
                diag[t] = (diag.get(t,"") + "|ticker_hist_ok").strip("|")
                continue
            fast = getattr(tk, "fast_info", None)
            if fast and getattr(fast, "previous_close", None):
                res[t] = (float(getattr(fast, "previous_close")), "yf-previousClose")
                diag[t] = (diag.get(t,"") + "|prevclose_ok").strip("|")
                continue
            info = getattr(tk, "info", {}) or {}
            if "previousClose" in info and info["previousClose"]:
                res[t] = (float(info["previousClose"]), "yf-info-prevClose")
                diag[t] = (diag.get(t,"") + "|info_prevclose_ok").strip("|")
                continue
            diag[t] = (diag.get(t,"") + "|ultra_fallback_fail").strip("|")
        except Exception as e:
            diag[t] = (diag.get(t,"") + f"|ultra_fallback_error:{e}").strip("|")
    return res, diag

# ============== Intraday orquestador ==============
def is_market_open_mx(now: Optional[datetime] = None) -> bool:
    # BMV: Lun–Vie, aprox 08:30–15:00 hora CDMX (simple, sin feriados)
    if now is None: now = datetime.now()
    # Nota: asume zona local ya en hora de CDMX
    if now.weekday() >= 5:  # 5=Sat, 6=Sun
        return False
    open_t = now.replace(hour=8, minute=30, second=0, microsecond=0)
    close_t = now.replace(hour=15, minute=0, second=0, microsecond=0)
    return open_t <= now <= close_t

def fetch_prices_live_or_daily(tickers_mx: List[str], use_intraday: bool, interval: str, window_min: int, prefer_daily: str):
    """
    Si use_intraday y mercado abierto: intenta intraday primero, luego cae al diario.
    Si mercado cerrado o falla intraday: usa diario directamente.
    """
    if not tickers_mx:
        return {}, {}

    diag_total: Dict[str, str] = {}
    res_total: Dict[str, Tuple[float, str]] = {}

    if use_intraday and is_market_open_mx():
        intra_map, intra_diag = yf_intraday_dict(tickers_mx, interval=interval, window_min=window_min)
        for k,v in intra_diag.items():
            diag_total[k] = (diag_total.get(k,"") + "|" + v).strip("|")
        res_total.update(intra_map)

        need = [t for t in tickers_mx if t not in res_total]
        if need:
            daily_map, daily_diag = fetch_prices_multi(need, prefer=prefer_daily)
            res_total.update(daily_map)
            for k,v in daily_diag.items():
                diag_total[k] = (diag_total.get(k,"") + "|" + v).strip("|")
    else:
        daily_map, daily_diag = fetch_prices_multi(tickers_mx, prefer=prefer_daily)
        res_total.update(daily_map)
        for k,v in daily_diag.items():
            diag_total[k] = (diag_total.get(k,"") + "|" + v).strip("|")

    return res_total, diag_total

# ================== ATR & Targets ==================
def compute_atr_14(ticker_mx: str, lookback_days: int = 60) -> Optional[float]:
    try:
        df = yf.download(tickers=ticker_mx, period=f"{lookback_days}d", interval="1d",
                         progress=False, group_by="ticker", threads=True, auto_adjust=False)
        if df is None or df.empty:
            return None
        if isinstance(df.columns, pd.MultiIndex):
            try:
                df = df[ticker_mx]
            except Exception:
                pass
        high = df["High"].astype(float)
        low  = df["Low"].astype(float)
        close = df["Close"].astype(float)

        prev_close = close.shift(1)
        tr1 = (high - low).abs()
        tr2 = (high - prev_close).abs()
        tr3 = (low - prev_close).abs()
        tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
        atr = tr.ewm(alpha=1/14, adjust=False).mean()
        val = float(atr.dropna().iloc[-1]) if not atr.dropna().empty else None
        return val
    except Exception as e:
        print(f"[ATR] Error {ticker_mx}: {e}")
        return None

def compute_tp_sl_buy_hybrid(entry: float, atr: Optional[float], atr_mult: float,
                             fallback_sl_pct: float, min_rr: float, tp_pct_policy: Optional[float],
                             tp_ceiling_atr_mult: Optional[float] = None):
    fallback_sl_abs = entry * float(fallback_sl_pct or 0.0)
    atr_sl_abs = (atr or 0.0) * float(atr_mult or 0.0)
    risk_per_share = max(fallback_sl_abs, atr_sl_abs)
    sl = max(0.01, entry - risk_per_share)

    tp_rr = entry + float(min_rr) * (entry - sl)
    tp_pct_abs = entry * (1.0 + float(tp_pct_policy or 0.0))
    tp = max(tp_rr, tp_pct_abs)

    if tp_ceiling_atr_mult is not None and atr is not None:
        tp_ceiling = entry + float(tp_ceiling_atr_mult) * atr
        tp = min(tp, tp_ceiling)

    risk_per_share = entry - sl
    rr = (tp - entry) / max(1e-9, risk_per_share)
    tp_pct_vs_entry = (tp - entry) / entry
    sl_pct_vs_entry = (entry - sl) / entry
    return sl, tp, risk_per_share, rr, tp_pct_vs_entry, sl_pct_vs_entry

# ================== Broker ==================
class PaperBroker:
    def __init__(self, per_trade_cash: float, max_open: int, budget: float,
                 state_file="active_positions.json", excel_path="trading_log.xlsx"):
        self.per_trade_cash = per_trade_cash
        self.max_open = max_open
        self.budget = budget
        self.state_file = state_file
        self.excel_path = excel_path
        self._load()

        ensure_excel_with_headers(self.excel_path, "opens", [
            "opened_at","opened_date","opened_time","opened_minute",
            "id","ticker","ticker_yf","side",
            "entry_price","qty","investment","sl","sl_pct","tp","tp_pct","rr",
            "prob","atr_used","cash_available_post_open","month","price_source"
        ])
        ensure_excel_with_headers(self.excel_path, "closes", [
            "closed_at","closed_date","closed_time","closed_minute",
            "id","ticker","ticker_yf","side",
            "entry_price","exit_price","qty","investment",
            "pnl","pnl_pct","result","held_days",
            "sl","tp","rr","prob","atr_used","month","price_source"
        ])
        ensure_excel_with_headers(self.excel_path, "skips", [
            "ts","date","time","minute","month",
            "ticker","ticker_yf","side","reason_code","reason_msg",
            "rr","min_rr","atr_used","sl_pct","tp_pct","entry_price",
            "price_source","per_trade_cash","budget_used","budget_total","max_open","active_count","price_diag"
        ])
        ensure_excel_with_headers(self.excel_path, "forecast_dates", [
            "ts","used_dates","n_dates","n_buy_considered","n_total_considered","month"
        ])

    def _load(self):
        if os.path.exists(self.state_file):
            with open(self.state_file, "r", encoding="utf-8") as f:
                self.positions = json.load(f)
            changed = False
            for pid, p in list(self.positions.items()):
                if "ticker_yf" not in p and "ticker" in p:
                    p["ticker_yf"] = ensure_mx(p["ticker"]); changed = True
                if "rr" not in p: p["rr"] = None; changed = True
                if "prob" not in p: p["prob"] = None; changed = True
                if "sl_pct" not in p and "entry_price" in p and "sl" in p:
                    try: p["sl_pct"] = max(0.0, (p["entry_price"] - p["sl"]) / p["entry_price"])
                    except: p["sl_pct"] = None
                    changed = True
                if "tp_pct" not in p and "entry_price" in p and "tp" in p:
                    try: p["tp_pct"] = max(0.0, (p["tp"] - p["entry_price"]) / p["entry_price"])
                    except: p["tp_pct"] = None
                    changed = True
                if "atr_used" not in p: p["atr_used"] = None; changed = True
                if "month" not in p: p["month"] = ""; changed = True
                if "price_source" not in p: p["price_source"] = None; changed = True
            if changed:
                with open(self.state_file, "w", encoding="utf-8") as f:
                    json.dump(self.positions, f, indent=2, ensure_ascii=False)
        else:
            self.positions = {}

    def _save(self):
        with open(self.state_file, "w", encoding="utf-8") as f:
            json.dump(self.positions, f, indent=2, ensure_ascii=False)

    def active_count(self):
        return sum(1 for p in self.positions.values() if p["status"] == "ACTIVE")

    def used_budget(self):
        return sum(p["investment"] for p in self.positions.values() if p["status"] == "ACTIVE")

    @staticmethod
    def _split_dt_fields(dt_iso: str):
        dt = datetime.fromisoformat(dt_iso)
        return dt.date().isoformat(), dt.strftime("%H:%M:%S"), f"{dt.minute:02d}"

    def log_open_excel(self, p: dict):
        opened_date, opened_time, opened_minute = self._split_dt_fields(p["opened_at"])
        append_excel_row(self.excel_path, "opens", [
            p["opened_at"], opened_date, opened_time, opened_minute,
            p["id"], p["ticker"], get_ticker_yf_from_pos(p), p["side"],
            p["entry_price"], p["qty"], p["investment"], p["sl"], p["sl_pct"], p["tp"], p["tp_pct"], p["rr"],
            p.get("prob", None), p.get("atr_used", None), p.get("cash_available_post_open", None), p.get("month",""), p.get("price_source")
        ])

    def log_close_excel(self, p: dict, pnl: float, pnl_pct: float, held_days: int):
        closed_date, closed_time, closed_minute = self._split_dt_fields(p["exit_at"])
        append_excel_row(self.excel_path, "closes", [
            p["exit_at"], closed_date, closed_time, closed_minute,
            p["id"], p["ticker"], get_ticker_yf_from_pos(p), p["side"],
            p["entry_price"], p["exit_price"], p["qty"], p["investment"],
            pnl, pnl_pct, p.get("result",""), held_days,
            p["sl"], p["tp"], p.get("rr", None), p.get("prob", None), p.get("atr_used", None), p.get("month",""), p.get("price_source")
        ])

    def log_skip_excel(self, row: dict):
        append_excel_row(self.excel_path, "skips", [
            row.get("ts"), row.get("date"), row.get("time"), row.get("minute"), row.get("month"),
            row.get("ticker"), row.get("ticker_yf"), row.get("side"), row.get("reason_code"), row.get("reason_msg"),
            row.get("rr"), row.get("min_rr"), row.get("atr_used"), row.get("sl_pct"), row.get("tp_pct"),
            row.get("entry_price"), row.get("price_source"), row.get("per_trade_cash"), row.get("budget_used"),
            row.get("budget_total"), row.get("max_open"), row.get("active_count"), row.get("price_diag")
        ])

    def log_forecast_dates_excel(self, used_dates: List[date], n_buy: int, n_total: int, month: str):
        ts = datetime.now().isoformat()
        append_excel_row(self.excel_path, "forecast_dates", [
            ts,
            ",".join([d.isoformat() for d in used_dates]),
            len(used_dates), n_buy, n_total, month
        ])

    def try_open_buy(self, ticker: str, price: float, sl: float, tp: float, horizon_days: int,
                     per_trade_cash: float, prob, rr: float, month: str,
                     sl_pct: float, tp_pct: float, atr_used: float, ticker_yf: str, price_source: str):
        if self.active_count() >= self.max_open:
            return None, "MAX_OPEN_REACHED"
        if self.used_budget() + per_trade_cash > self.budget:
            return None, "BUDGET_EXCEEDED"

        qty = max(1, int(per_trade_cash // price))
        if qty <= 0:
            return None, "QTY_ZERO"
        invest = qty * price

        pid = f"{ticker}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.positions[pid] = {
            "id": pid,
            "ticker": ticker,
            "ticker_yf": ticker_yf,
            "side": "BUY",
            "entry_price": price,
            "tp": tp,
            "sl": sl,
            "sl_pct": sl_pct,
            "tp_pct": tp_pct,
            "qty": qty,
            "investment": invest,
            "opened_at": datetime.now().isoformat(),
            "max_holding_days": int(horizon_days),
            "status": "ACTIVE",
            "prob": prob,
            "rr": rr,
            "atr_used": atr_used,
            "month": month,
            "price_source": price_source
        }
        self._save()
        self.positions[pid]["cash_available_post_open"] = self.budget - self.used_budget()
        self._save()
        self.log_open_excel(self.positions[pid])
        return pid, "OPENED"

    def _force_close(self, pid: str, last_price: float, reason: str):
        p = self.positions[pid]
        qty = p["qty"]; entry = p["entry_price"]
        pnl = (last_price - entry) * qty
        p["status"] = "CLOSED"
        p["exit_price"] = last_price
        p["exit_at"] = datetime.now().isoformat()
        p["pnl"] = pnl
        p["result"] = reason
        self._save()
        held_days = (datetime.fromisoformat(p["exit_at"]) - datetime.fromisoformat(p["opened_at"])).days
        pnl_pct = pnl / max(1e-9, p["investment"])
        self.log_close_excel(p, pnl, pnl_pct, held_days)
        return pnl, pnl_pct, held_days

    def try_close_hits(self, pid: str, last_price: float):
        p = self.positions[pid]
        if p["status"] != "ACTIVE": return False, None, 0.0, 0.0, 0
        tp = p["tp"]; sl = p["sl"]; entry = p["entry_price"]
        hit = None
        if last_price >= tp: hit = "TP_HIT"
        elif last_price <= sl: hit = "SL_HIT"

        if hit:
            pnl, pnl_pct, held_days = self._force_close(pid, last_price, hit)
            return True, hit, pnl, pnl_pct, held_days

        opened = datetime.fromisoformat(p["opened_at"])
        if datetime.now() >= opened + timedelta(days=int(p["max_holding_days"])):
            pnl, pnl_pct, held_days = self._force_close(pid, last_price, "TIME_EXIT")
            return True, "TIME_EXIT", pnl, pnl_pct, held_days

        return False, None, 0.0, 0.0, 0

# ================== Cierre masivo ==================
def close_all_active_now(broker: PaperBroker):
    active = [(pid, p) for pid, p in broker.positions.items() if p["status"] == "ACTIVE"]
    if not active:
        print("No hay posiciones activas para cerrar.")
        return 0, 0.0

    tickers = list({ensure_mx(p["ticker"]) for _, p in active})
    prices = {}
    price_map, _ = fetch_prices_live_or_daily(tickers, use_intraday=True, interval="1m", window_min=10, prefer_daily="yf")
    for t in tickers:
        if t in price_map:
            prices[t] = price_map[t][0]

    closed_n = 0
    total_pnl = 0.0
    for pid, p in active:
        tmx = ensure_mx(p["ticker"])
        if tmx not in prices:
            print(f"[RESET] Sin precio para {tmx}, salto cierre forzado.")
            continue
        pnl, pnl_pct, held_days = broker._force_close(pid, prices[tmx], "MANUAL_RESET")
        closed_n += 1
        total_pnl += pnl
        rr_txt = f"{p.get('rr',0):.2f}x" if p.get("rr") is not None else "-"
        prob_txt = prob_to_pct(p.get("prob", None))
        send_tg(
            f"⚠️ <b>CIERRE MASIVO</b>\n"
            f"📈 <b>{p['ticker']}</b> ({get_ticker_yf_from_pos(p)}) BUY  |  Qty: <b>{p['qty']}</b>\n"
            f"📍 MANUAL_RESET  |  ⏳ {held_days}d  |  R≈{rr_txt}\n"
            f"💰 Entrada: {pretty_cash(p['entry_price'])}  →  Salida: {pretty_cash(prices[tmx])}\n"
            f"📊 P&L: <b>{pretty_cash(pnl)}</b> ({pnl_pct*100:+.1f}%)\n"
            f"🔮 Prob. éxito (señal): {prob_txt}"
        )
    print(f"[RESET] Cerradas {closed_n} posiciones. PnL total ~ {pretty_cash(total_pnl)}")
    return closed_n, total_pnl

# ================== MAIN ==================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--month", required=True, help="Mes objetivo YYYY-MM (ej. 2025-10)")
    parser.add_argument("--base-dir", default=".", help="Raíz del repo")
    parser.add_argument("--interval-min", type=int, default=10, help="Minutos entre chequeos")
    parser.add_argument("--max-new", type=int, default=3, help="Máximo de aperturas por ciclo")
    parser.add_argument("--excel-path", type=str, default="trading_log.xlsx", help="Excel de bitácora")
    parser.add_argument("--positions-file", type=str, default="open_positions.json", help="Archivo de posiciones abiertas")

    # Solo BUY + control dinámico
    parser.add_argument("--only-buy", action="store_true", default=True, help="Forzar solo BUY")
    parser.add_argument("--atr-mult", type=float, default=4.0, help="Multiplicador ATR para SL")
    parser.add_argument("--min-rr", type=float, default=4.0, help="Relación mínima R:R")
    parser.add_argument("--fallback-sl-pct", type=float, default=0.001, help="SL % respaldo si no hay ATR (0.1%)")
    parser.add_argument("--tp-ceiling-atr-mult", type=float, default=5.0, help="Techo TP = entry + N*ATR")

    # Fechas recientes del forecast
    parser.add_argument("--recent-days", type=int, default=3, help="N fechas recientes del forecast a considerar")

    # Preferencia y live
    parser.add_argument("--prefer", type=str, default="yf", choices=["yf","investing"], help="Fuente preferida diaria")
    parser.add_argument("--use-intraday", action="store_true", default=True, help="Intentar intraday cuando el mercado esté abierto")
    parser.add_argument("--intraday-interval", type=str, default="1m", choices=["1m","5m"], help="Intervalo intradía YF")
    parser.add_argument("--intraday-window-min", type=int, default=15, help="Ventana (min) para tomar último precio válido")

    # Reset al iniciar
    parser.add_argument("--ask-reset", action="store_true", default=True, help="Preguntar si cerrar todas las posiciones al inicio")
    parser.add_argument("--force-reset", action="store_true", default=False, help="Cerrar todas sin preguntar")
    parser.add_argument("--no-ask-reset", action="store_true", default=False, help="No preguntar por reset al inicio")

    # Skips / debug
    parser.add_argument("--log-all-candidates", action="store_true", default=False, help="Loggear todos los candidatos en skips")
    parser.add_argument("--skips-to-telegram", action="store_true", default=False, help="Enviar resumen de skips a Telegram")

    args = parser.parse_args()

    month = args.month
    base_dir = os.path.abspath(args.base_dir)

    # Política
    policy = load_policy(month, base_dir)
    tp_pct_base = float(policy["tp_pct"] or 0.0)
    sl_pct_fallback = float(policy["sl_pct"] or args.fallback_sl_pct)
    horizon_days = int(policy["horizon_days"])
    per_trade_cash = float(policy["per_trade_cash"])
    max_open = int(policy["max_open"])
    budget = float(policy["budget"])

    print("=== Paper monitor BMV (BUY, ATR SL, TP híbrido, Excel, Live v7) ===")
    print(f"Mes: {month}")
    print(f"H={horizon_days}d  cash/trade={pretty_cash(per_trade_cash)}  max_open={max_open}  budget={pretty_cash(budget)}")
    print(f"SL: ATR*{args.atr_mult}  (respaldo {sl_pct_fallback*100:.2f}%) | RR≥{args.min_rr:.1f}x | tp_pct={tp_pct_base*100:.1f}% | Techo={args.tp_ceiling_atr_mult}*ATR")
    print(f"Excel: {args.excel_path} | Positions: {args.positions_file} | Fechas recientes: {args.recent_days}")
    print(f"Live: use_intraday={args.use_intraday} interval={args.intraday_interval} window={args.intraday_window_min}m | Prefer diaria: {args.prefer}")

    send_tg(
        f"🚀 <b>MONITOR BMV</b> Live v7 (BUY, SL=ATR*{args.atr_mult}, RR≥{args.min_rr}x, TP híbrido)\n"
        f"🗓 {month} | ⏳ H {horizon_days}d | 💵 {pretty_cash(per_trade_cash)} x trade | 📊 max_open={max_open} | 💼 {pretty_cash(budget)}\n"
        f"📡 Intradía: {'ON' if args.use_intraday else 'OFF'} ({args.intraday_interval}, {args.intraday_window_min}m) · Prefer diaria: {args.prefer.upper()}"
    )

    # Broker
    broker = PaperBroker(
        per_trade_cash=per_trade_cash,
        max_open=max_open,
        budget=budget,
        state_file=args.positions_file,
        excel_path=args.excel_path
    )

    # Reset inicial
    if args.force_reset:
        print("[RESET] Forzado por flag --force-reset.")
        close_all_active_now(broker)
    elif not args.no_ask_reset and args.ask_reset:
        if broker.active_count() > 0:
            try:
                resp = input("¿Cerrar TODAS las posiciones activas ahora para reiniciar? [s/N]: ").strip().lower()
            except EOFError:
                resp = "n"
            if resp in ("s","si","sí","y","yes"):
                close_all_active_now(broker)
            else:
                print("[RESET] No se cerraron posiciones al inicio.")

    # Forecast -> últimas N fechas
    df_fc = load_forecast(month, base_dir)
    today = datetime.now().date()
    cand_df, used_dates = pick_recent_signals(df_fc, today, n_recent_days=args.recent_days, only_buy=args.only_buy)

    # Log fechas consideradas
    broker.log_forecast_dates_excel(used_dates, n_buy=len(cand_df), n_total=len(cand_df), month=month)
    print(f"📊 Usando fechas: {[d.isoformat() for d in used_dates]} | BUY considerados: {len(cand_df)}")
    if cand_df.empty:
        print("No hay señales BUY en el rango de fechas considerado.")
        send_tg("ℹ️ No hay señales BUY en el forecast (en fechas recientes).")
        return

    print("Candidatos BUY (mejor score por ticker):")
    print(cand_df.head(20).to_string(index=False))

    # ===== Loop =====
    while True:
        try:
            # 1) Cierre por TP/SL/horizonte
            active = [(pid, p) for pid, p in broker.positions.items() if p["status"] == "ACTIVE"]
            if active:
                print(f"[{datetime.now():%H:%M:%S}] Activas: {len(active)}")
                tickers = list({ensure_mx(p["ticker"]) for _, p in active})
                price_map, price_diag = fetch_prices_live_or_daily(
                    tickers, use_intraday=args.use_intraday,
                    interval=args.intraday_interval, window_min=args.intraday_window_min,
                    prefer_daily=args.prefer
                )
                for pid, p in active:
                    tmx = ensure_mx(p["ticker"])
                    if tmx in price_map:
                        last_px, _src = price_map[tmx]
                        closed, reason, pnl, pnl_pct, held_days = broker.try_close_hits(pid, last_px)
                        if closed:
                            emoji = "🎉" if reason == "TP_HIT" else ("🛑" if reason == "SL_HIT" else "⏰")
                            entry = p["entry_price"]; exit_px = broker.positions[pid]["exit_price"]
                            qty = p["qty"]; rr_txt = f"{p.get('rr',0):.2f}x" if p.get("rr") is not None else "-"
                            prob_txt = prob_to_pct(p.get("prob", None))
                            send_tg(
                                f"{emoji} <b>POSICIÓN CERRADA</b>\n"
                                f"📈 <b>{p['ticker']}</b> ({get_ticker_yf_from_pos(p)}) BUY  |  Qty: <b>{qty}</b>\n"
                                f"📍 {reason}  |  ⏳ {held_days}d  |  R≈{rr_txt}\n"
                                f"💰 Entrada: {pretty_cash(entry)}  →  Salida: {pretty_cash(exit_px)}\n"
                                f"📊 P&L: <b>{pretty_cash(pnl)}</b> ({pnl_pct*100:+.1f}%)\n"
                                f"🔮 Prob. éxito: {prob_txt}"
                            )
                    else:
                        # Log skip por falta de precio de cierre
                        now = datetime.now()
                        broker.log_skip_excel({
                            "ts": now.isoformat(), "date": now.date().isoformat(),
                            "time": now.strftime("%H:%M:%S"), "minute": f"{now.minute:02d}", "month": month,
                            "ticker": p["ticker"], "ticker_yf": ensure_mx(p["ticker"]), "side": "BUY",
                            "reason_code": "no_price_close",
                            "reason_msg": "No hay precio para evaluar cierre",
                            "rr": p.get("rr"), "min_rr": None, "atr_used": p.get("atr_used"),
                            "sl_pct": p.get("sl_pct"), "tp_pct": p.get("tp_pct"), "entry_price": p.get("entry_price"),
                            "price_source": None, "per_trade_cash": per_trade_cash,
                            "budget_used": broker.used_budget(), "budget_total": broker.budget,
                            "max_open": broker.max_open, "active_count": broker.active_count(),
                            "price_diag": price_diag.get(tmx, "")
                        })

            # 2) Apertura nuevas
            capacity = max(0, broker.max_open - broker.active_count())
            to_open = min(capacity, args.max_new)
            if to_open > 0:
                active_tickers = {p["ticker"] for p in broker.positions.values() if p["status"] == "ACTIVE"}
                open_candidates = cand_df[~cand_df["ticker"].isin(active_tickers)].head(to_open).copy()
                tickers_new_mx = [ensure_mx(t) for t in open_candidates["ticker"].tolist()]

                price_map, price_diag = fetch_prices_live_or_daily(
                    tickers_new_mx, use_intraday=args.use_intraday,
                    interval=args.intraday_interval, window_min=args.intraday_window_min,
                    prefer_daily=args.prefer
                )

                for _, row in open_candidates.iterrows():
                    t = str(row["ticker"])
                    tmx = ensure_mx(t)
                    if tmx not in price_map:
                        now = datetime.now()
                        broker.log_skip_excel({
                            "ts": now.isoformat(), "date": now.date().isoformat(),
                            "time": now.strftime("%H:%M:%S"), "minute": f"{now.minute:02d}", "month": month,
                            "ticker": t, "ticker_yf": tmx, "side": "BUY",
                            "reason_code": "no_price_open",
                            "reason_msg": "No hay precio para abrir",
                            "rr": None, "min_rr": args.min_rr, "atr_used": None,
                            "sl_pct": None, "tp_pct": None, "entry_price": None,
                            "price_source": None, "per_trade_cash": per_trade_cash,
                            "budget_used": broker.used_budget(), "budget_total": broker.budget,
                            "max_open": broker.max_open, "active_count": broker.active_count(),
                            "price_diag": price_diag.get(tmx, "")
                        })
                        continue

                    price, price_source = price_map[tmx]
                    atr_val = compute_atr_14(tmx, lookback_days=60)
                    sl, tp, risk_ps, rr, tp_pct_vs_entry, sl_pct_vs_entry = compute_tp_sl_buy_hybrid(
                        entry=price, atr=atr_val, atr_mult=args.atr_mult,
                        fallback_sl_pct=sl_pct_fallback, min_rr=args.min_rr,
                        tp_pct_policy=tp_pct_base, tp_ceiling_atr_mult=args.tp_ceiling_atr_mult
                    )

                    if rr < args.min_rr:
                        now = datetime.now()
                        broker.log_skip_excel({
                            "ts": now.isoformat(), "date": now.date().isoformat(),
                            "time": now.strftime("%H:%M:%S"), "minute": f"{now.minute:02d}", "month": month,
                            "ticker": t, "ticker_yf": tmx, "side": "BUY",
                            "reason_code": "rr_below_min",
                            "reason_msg": f"R {rr:.2f} < {args.min_rr:.2f}",
                            "rr": rr, "min_rr": args.min_rr, "atr_used": atr_val,
                            "sl_pct": sl_pct_vs_entry, "tp_pct": tp_pct_vs_entry, "entry_price": price,
                            "price_source": price_source, "per_trade_cash": per_trade_cash,
                            "budget_used": broker.used_budget(), "budget_total": broker.budget,
                            "max_open": broker.max_open, "active_count": broker.active_count(),
                            "price_diag": price_diag.get(tmx, "")
                        })
                        continue

                    pid, status = broker.try_open_buy(
                        ticker=t, price=price, sl=sl, tp=tp, horizon_days=horizon_days,
                        per_trade_cash=per_trade_cash, prob=float(row.get("score", 0.0)),
                        rr=rr, month=month, sl_pct=sl_pct_vs_entry, tp_pct=tp_pct_vs_entry,
                        atr_used=float(atr_val) if atr_val is not None else None,
                        ticker_yf=tmx, price_source=price_source
                    )

                    if status == "OPENED":
                        p = broker.positions[pid]
                        qty = p["qty"]; invest = p["investment"]
                        pnl_target = (tp - price) * qty
                        prob_txt = prob_to_pct(p.get("prob", None))
                        send_tg(
                            f"🚀 <b>NUEVA POSICIÓN</b>\n"
                            f"📈 <b>{t}</b> ({tmx}) BUY  |  Fuente: <b>{price_source}</b>\n"
                            f"💰 Entrada: {pretty_cash(price)}  |  Qty: <b>{qty}</b>\n"
                            f"🧮 Inversión: {pretty_cash(invest)}  |  💼 Cash post: {pretty_cash(p.get('cash_available_post_open', 0.0))}\n"
                            f"🛡️ SL: {pretty_cash(sl)}  ({sl_pct_vs_entry*100:.2f}%)\n"
                            f"🎯 TP: {pretty_cash(tp)}  ({tp_pct_vs_entry*100:.2f}%)\n"
                            f"📏 ATR usado: {pretty_cash(p.get('atr_used', 0.0) or 0.0)}\n"
                            f"🔢 R≈{rr:.2f}x  |  🔮 Prob. éxito: {prob_txt}\n"
                            f"📊 Ganancia estimada a TP: <b>{pretty_cash(pnl_target)}</b>\n"
                            f"⏳ Horizonte: {horizon_days}d"
                        )
                    else:
                        now = datetime.now()
                        broker.log_skip_excel({
                            "ts": now.isoformat(), "date": now.date().isoformat(),
                            "time": now.strftime("%H:%M:%S"), "minute": f"{now.minute:02d}", "month": month,
                            "ticker": t, "ticker_yf": tmx, "side": "BUY",
                            "reason_code": status.lower(), "reason_msg": f"No se abrió: {status}",
                            "rr": rr, "min_rr": args.min_rr, "atr_used": atr_val,
                            "sl_pct": sl_pct_vs_entry, "tp_pct": tp_pct_vs_entry, "entry_price": price,
                            "price_source": price_source, "per_trade_cash": per_trade_cash,
                            "budget_used": broker.used_budget(), "budget_total": broker.budget,
                            "max_open": broker.max_open, "active_count": broker.active_count(),
                            "price_diag": price_diag.get(tmx, "")
                        })

            # Opcional: log de todos los candidatos vistos
            if args.log_all_candidates:
                now = datetime.now()
                used_ts = now.isoformat()
                for _, row in cand_df.iterrows():
                    t = str(row["ticker"]); tmx = ensure_mx(t)
                    append_excel_row(broker.excel_path, "skips", [
                        used_ts, now.date().isoformat(), now.strftime("%H:%M:%S"), f"{now.minute:02d}", month,
                        t, tmx, "BUY", "candidate_seen", "Candidato evaluado (tracking)",
                        None, args.min_rr, None, None, None, None,
                        None, per_trade_cash, broker.used_budget(), broker.budget, broker.max_open, broker.active_count(), ""
                    ])

            # 3) Espera
            wake = (datetime.now()+timedelta(minutes=args.interval_min)).strftime("%H:%M")
            print(f"Duerme {args.interval_min} min - {wake}")
            time.sleep(args.interval_min * 60)

        except KeyboardInterrupt:
            print("Detenido por usuario.")
            send_tg("🛑 Monitor BMV detenido por el usuario.")
            break
        except Exception as e:
            print(f"Error ciclo: {e}")
            time.sleep(10)

if __name__ == "__main__":
    main()





v